# -*- coding: utf-8 -*-

###############################################################################
#                              hom_calculos.py                                #
###############################################################################

import ctypes
import pyautogui
import logging
import time
import os
import base64
from anthropic import Anthropic
from openai import OpenAI
import pyexcel_xls
from pywinauto import Application
from pywinauto.findwindows import ElementNotFoundError, find_windows
from pywinauto.timings import TimeoutError
import cv2
from pywinauto import Desktop
from datetime import date, timedelta
import calendar
from holidays import Brazil
import openpyxl
import unicodedata, re, difflib
import shutil
import psutil
import signal
from dotenv import load_dotenv
import tempfile
import sys

pyautogui.FAILSAFE = False
pyautogui.PAUSE = 0.1

# =====================================================================
# üîß Carrega .env a partir da MESMA pasta do .exe (igual conc_shopping)
# =====================================================================
base_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(base_dir, ".env")

if os.path.exists(env_path):
    load_dotenv(dotenv_path=env_path)
    print(f"[DEBUG] .env carregado de: {env_path}")
else:
    print(f"[ERRO] .env n√£o encontrado em: {env_path}")

# =====================================================================
# üîë API Keys carregadas do .env (sem expor no c√≥digo)
# =====================================================================
openai_api_key = os.getenv("OPENAI_API_KEY")
anthropic_api_key = os.getenv("ANTHROPIC_API_KEY")

if not openai_api_key:
    print("[ERRO] OPENAI_API_KEY n√£o encontrada no .env")
if not anthropic_api_key:
    print("[ERRO] ANTHROPIC_API_KEY n√£o encontrada no .env")

client = OpenAI(api_key=openai_api_key)
anthropic = Anthropic(api_key=anthropic_api_key)

# =====================================================================
# üìÅ Diret√≥rios globais fixos como no conc_shopping
# =====================================================================

root_dir = r"C:\AUTOMACAO\faturamento"

log_dir = os.path.join(root_dir, "Logs")
prints_dir = os.path.join(root_dir, "prints")

os.makedirs(log_dir, exist_ok=True)
os.makedirs(prints_dir, exist_ok=True)

print(f"[DEBUG] Logs fixos em: {log_dir}")
print(f"[DEBUG] Prints fixos em: {prints_dir}")

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt='%d/%m/%Y %H:%M:%S'
)

br_holidays = Brazil()

folder_map = {
    "Shopping Montserrat": r"C:\Program Files\Victor & Schellenberger\VSSC_MONTSERRAT",
    "Shopping da Ilha": r"C:\Program Files\Victor & Schellenberger\VSSC_ILHA",
    "Shopping Rio Poty": r"C:\Program Files\Victor & Schellenberger\VSSC_TERESINA",
    "Shopping Metr√≥pole": r"C:\Program Files\Victor & Schellenberger\VSSC_METROPOLE",
    "Shopping Moxuara": r"C:\Program Files\Victor & Schellenberger\VSSC_MOXUARA",
    "Shopping Praia da Costa": r"C:\Program Files\Victor & Schellenberger\VSSC_PRAIADACOSTA",
    "Shopping Mestre √Ålvaro": r"C:\Program Files\Victor & Schellenberger\VSSC_MESTREALVARO"
}

shopping_fases_tipo2 = {
    "Shopping Mestre √Ålvaro": {
        "Antecipados": [5, 7, 24, 25],
        "At√≠picos": [31, 32, 6, 4, 30],
        "Postecipados": [8, 11, 2, 41]
    },
    "Shopping Montserrat": {
        "Antecipados": [5, 7, 24, 25],
        "At√≠picos": [31, 32, 6, 11, 30],
        "Postecipados": [22, 11, 4, 22]
    },
    "Shopping Metr√≥pole": {
        "Antecipados": [7, 18, 12, 13],
        "At√≠picos": [31, 32, 6, 11],
        "Postecipados": [2, 8, 9, 36]
    },
    "Shopping Praia da Costa": {
        "Antecipados": [3, 20, 4, 5],
        "At√≠picos": [31, 32, 6, 24],
        "Postecipados": [8, 11, 13]
    },
    "Shopping Rio Poty": {
        "Antecipados": [7, 18, 12, 13],
        "At√≠picos": [31, 32, 6, 11],
        "Postecipados": [2, 8, 9, 23]
    },
    "Shopping da Ilha": {
        "Antecipados": [7, 18, 12, 13],
        "At√≠picos": [31, 32, 6, 11, 30],
        "Postecipados": [2, 8, 9, 36, 37]
    },
    "Shopping Moxuara": {
        "Antecipados": [24, 18, 12, 13],
        "At√≠picos": [31, 32, 6, 11, 30],
        "Postecipados": [2, 8, 9, 39]
    }
}

# Substitu√≠mos todas as men√ß√µes a "Shopping Teresina" por "Shopping Rio Poty"
# para evitar a duplicidade e confus√£o entre Teresina e Rio Poty.

missing_phases_map = {
    "Shopping Montserrat": [29, 39, 40, 41, 42, 43, 44, 45],
    "Shopping da Ilha": [3],
    "Shopping Mestre √Ålvaro": [12, 13, 38, 43, 44, 46, 47, 48, 49],
    "Shopping Metr√≥pole": [3, 23, 29],
    "Shopping Moxuara": [],
    "Shopping Praia da Costa": [27, 42],
    "Shopping Rio Poty": [3, 39, 43, 44, 45, 46, 47, 48, 49]  
}
 
shopping_num_map = {
    "Shopping Mestre √Ålvaro": 2,
    "Shopping Montserrat": 3,
    "Shopping da Ilha": 1,
    "Shopping Metr√≥pole": 3,
    "Shopping Moxuara": 2,
    "Shopping Praia da Costa": 1,
    "Shopping Rio Poty": 1
}

prints_folder = os.path.join(os.getcwd(), "prints")
SCREENSHOT_PATH = os.path.join(prints_folder, "monitor_screenshot.png")  # Sempre substitui esse arquivo
if not os.path.exists(prints_folder):
    os.makedirs(prints_folder)

IS_SEGURO = False

def normalize(s: str) -> str:
    # 1) Desmonta acentos (NFKD) e joga fora marcas de combina√ß√£o
    nfkd = unicodedata.normalize("NFKD", s)
    without_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    # 2) Junta espa√ßos e quebras de linha em um √∫nico espa√ßo
    collapsed = re.sub(r"\s+", " ", without_accents)
    return collapsed.lower()

def fuzzy_contains(text: str, sub: str, threshold: float = 0.8) -> bool:
    text_norm = normalize(text)
    sub_norm  = normalize(sub)

    # 1) se vier ‚Äúaluguel minimo‚Äù literalmente
    if sub_norm in text_norm:
        return True

    # 2) janela deslizante no texto normalizado
    max_ratio = 0.0
    L = len(sub_norm)
    for i in range(len(text_norm) - L + 1):
        seg = text_norm[i : i+L]
        ratio = difflib.SequenceMatcher(None, seg, sub_norm).ratio()
        if ratio >= threshold:
            return True

    return False

def build_fase_map(shopping):
    """
    Constr√≥i um dicion√°rio que mapeia a 'fase' para a posi√ß√£o relativa de clique
    na tela, considerando as fases 'ausentes' (missing_phases_map).
    """
    base_y = 33
    step_y = 14
    coords = {}
    missing = missing_phases_map.get(shopping, [])
    real_index = 1
    for fase in range(1, 46):
        if fase in missing:
            continue
        if real_index < 14:
            coords[fase] = base_y + (real_index - 1) * step_y
        else:
            coords[fase] = 215
        real_index += 1
    return coords

def get_visible_index(shopping, fase):
    """
    Retorna quantas fases 'vis√≠veis' (i.e., n√£o ausentes em missing_phases_map)
    existem de 1 at√© 'fase' inclusive.
    """
    missing = missing_phases_map.get(shopping, [])
    index = 0
    for f in range(1, fase + 1):
        if f not in missing:
            index += 1
    return index


def click_fase_tipo1(shopping, fase):
    """
    Corrige o m√©todo de clicar na fase, respeitando as fases ausentes
    tamb√©m para fases acima de 13.
    """

    if not fase:
        logging.error(f"Fase inv√°lida para {shopping}. Verifique o mapeamento.")
        return

    try:
        fase_int = int(fase)
    except ValueError:
        logging.error(f"Fase n√£o num√©rica para {shopping}: {fase}")
        return

    # Formata com zero √† esquerda se necess√°rio
    fase_str = f"{fase_int:02d} -"

    # Registra no log o texto que ser√° digitado
    logging.info(f"[click_fase_tipo1] Digitando string de busca: '{fase_str}'")

    # Abre busca
    pyautogui.hotkey('ctrl', 'f')
    time.sleep(3)

    # Tenta garantir que a caixa de busca est√° ativa
    pyautogui.press('tab')
    time.sleep(0.5)
    

    # Digita com intervalo para aumentar a chance de entrada
    pyautogui.write(fase_str, interval=0.05)
    logging.info(f"[click_fase_tipo1] Teclas enviadas para busca: '{fase_str}'")
    time.sleep(2)
    

    pyautogui.press('enter')
    logging.info(f"Clicando na fase {fase}")
    time.sleep(3)
    pyautogui.press('down')
    time.sleep(10)
    pyautogui.hotkey('alt', 'space')
    time.sleep(0.3)
    pyautogui.press('down')
    time.sleep(0.3)
    pyautogui.press('enter')
    pyautogui.moveRel(95, 37)
    pyautogui.click()
    pyautogui.click()
    pyautogui.click()
    time.sleep(0.5)
    pyautogui.press('up')
    time.sleep(4)



def determine_variant(shopping):
    """
    Determina o variant com base no nome do shopping.
    """
    mapping = {
        "Shopping da Ilha": "SDI",
        "Shopping Mestre √Ålvaro": "SMA",
        "Shopping Moxuara": "SMO",
        "Shopping Montserrat": "SMS",
        "Shopping Metr√≥pole": "SMT",
        "Shopping Rio Poty": "SRP",
        "Shopping Praia da Costa": "SPC"
    }
    return mapping.get(shopping, "SDI")  # SDI como default se n√£o encontrar

def get_extra_phases(shopping, faturamento_mode, variant):
    """
    Retorna lista de contas/fases 'extras', que n√£o s√£o as contas tradicionais,
    mas sim varia√ß√µes espec√≠ficas, como postecipado de mall, espa√ßo anunciante etc.
    """
    extras = []
    # POSTECIPADO
    if faturamento_mode.upper() == "POSTECIPADO":
        if variant.upper() == "SDI":
            # AMM DANIEL / COND DANIEL
            extras.append(("200101", 36))  # Aluguel M√≠nimo - Daniel
            extras.append(("310100", 37)) 
            extras.append(("200101", 44)) # Encargo Comum - Daniel
        elif variant.upper() == "SMA":
            # Postecipado de MALL
            extras.append(("200101", 41))
        elif variant.upper() == "SMO":
            # Postecipado MALL
            extras.append(("200101", 39))
        elif variant.upper() == "SMS":
            extras.append(("200101", 22))
        elif variant.upper() == "SMT":
            extras.append(("200101", 36))
        elif variant.upper() == "SRP":
            extras.append(("200101", 23))
            ###FPPCELOJA ENCARGO FUNDO
            
    # ANTECIPADO
    elif faturamento_mode.upper() == "ANTECIPADO":
        if variant.upper() == "SDI":
            extras.append(("200119", 18))  # Espa√ßo Anunciante
            extras.append(("200156", 18))  # M√≠dia Inaugural
        elif variant.upper() == "SMA":
            extras.append(("200119", 7))   # Espa√ßo Anunciante
            extras.append(("200156", 7))   # M√≠dia Inaugural
        elif variant.upper() == "SMO":
            extras.append(("200119", 18))
            extras.append(("200156", 18))
        elif variant.upper() == "SMS":
            extras.append(("200119", 7))
            extras.append(("200156", 7))
        elif variant.upper() == "SMT":
            extras.append(("200119", 18))
            extras.append(("200156", 18))
        elif variant.upper() == "SRP":
            extras.append(("200119", 18))
            extras.append(("200156", 18))
    return extras

def get_phase(shopping, conta, faturamento_mode, variant):
    """
    Retorna a fase correta de acordo com (shopping -> variant), conta,
    e tipo de faturamento (postecipado ou antecipado).
    """
    phases = {
        "POSTECIPADO": {
            "SDI": {
                "310100": 8,
                "210100": 9,
                "200101": 2,
                "200106": 2,
                "410900": 8,
                "411000": 8,
            },
            "SMA": {
                "310100": 11,
                "210100": 2,
                "200101": 8,
                "200106": 8,
                "410900": 11,
                "411000": 11,
            },
            "SMO": {
                "310100": 8,
                "210100": 9,
                "200101": 2,
                "200106": 2,
                "410900": 8,
                "411000": 8,
            },
            "SMS": {
                "310100": 11,
                "210100": 4,
                "200101": 8,
                "200106": 8,
                "410900": 11,
                "411000": 11,
            },
            "SMT": {
                "310100": 8,
                "210100": 9,
                "200101": 2,
                "200106": 2,
                "410900": 8,
                "411000": 8,
            },
            "SRP": {
                "310100": 8,
                "210100": 9,
                "200101": 2,
                "200106": 2,
                "410900": 8,
                "411000": 8,
            }
        },
        "ANTECIPADO": {
            "SDI": {
                "310100": 12,
                "410900": 12,
                "411000": 12,
                "210100": 13,
                "200101": 7,
                "200133": 7,
            },
            "SMA": {
                "310100": 24,
                "410900": 24,
                "411000": 24,
                "210100": 25,
                "200101": 5,
                "200133": 5,
            },
            "SMO": {
                "310100": 12,
                "410900": 12,
                "411000": 12,
                "210100": 13,
                "200101": 24,
                "200133": 24,
            },
            "SMS": {
                "310100": 24,
                "410900": 24,
                "411000": 24,
                "210100": 25,
                "200101": 5,
                "200133": 5,
            },
            "SMT": {
                "310100": 12,
                "410900": 12,
                "411000": 12,
                "210100": 13,
                "200101": 7,
                "200133": 7,
            },
            "SRP": {
                "310100": 12,
                "410900": 12,
                "411000": 12,
                "210100": 13,
                "200101": 7,
                "200133": 7,
            }
        },
        "AT√çPICOS": {
            "SDI": {
                "311101": 31,
                "211101": 32,
                "300101": 6,
                "300101_mall": 11,
                "311128": 31,
                "311129": 31
            },
            "SMA": {
                "311101": 31,
                "211101": 32,
                "300101": 6,
                "300101_mall": 4,
                "311128": 31,
                "311129": 31
            },
            "SMO": {
                "311101": 31,
                "211101": 32,
                "300101": 6,
                "300101_mall": 11,
                "311128": 31,
                "311129": 31
            },
            "SMS": {
                "311101": 31,
                "211101": 32,
                "300101": 6,
                "300101_mall": 15,
                "311128": 31,
                "311129": 31
            },
            "SMT": {
                "311101": 31,
                "211101": 32,
                "300101": 6,
                "300101_mall": 11,
                "311128": 31,
                "311129": 31
            },
            "SRP": {
                "311101": 31,
                "211101": 32,
                "300101": 6,
                "300101_mall": 11,
                "311128": 31,
                "311129": 31
            }
        }
    }


    fm_upper = faturamento_mode.upper()
    var_upper = variant.upper()

    if fm_upper not in phases:
        return None
    if var_upper not in phases[fm_upper]:
        return None

    # Retorna a fase mapeada ou None se a conta n√£o estiver no dicion√°rio
    return phases[fm_upper][var_upper].get(conta)

def find_and_click_button_with_retry(image_path, max_attempts=10, confidence_range=(0.95, 0.6)):
    try:
        for attempt in range(max_attempts):
            confidence = confidence_range[0] - (attempt * (confidence_range[0] - confidence_range[1]) / max_attempts)
            logging.info("Tentativa %d/%d com confian√ßa %.2f" % (attempt + 1, max_attempts, confidence))
            try:
                button = pyautogui.locateOnScreen(image_path, confidence=confidence)
                if button:
                    x, y = pyautogui.center(button)
                    logging.info("Bot√£o encontrado nas coordenadas: x=%d, y=%d" % (x, y))
                    pyautogui.click(x, y)
                    return True
            except Exception as e:
                logging.error("Erro na tentativa %d: %s" % (attempt + 1, str(e)))
            time.sleep(2)
        logging.error("Bot√£o n√£o encontrado ap√≥s todas as tentativas")
        return False
    except Exception as e:
        logging.error("Erro ao tentar localizar bot√£o: %s" % str(e))
        return False

def verify_image_visibility(image_path, confidence=0.7, max_retries=10):
    try:
        logging.info("Verificando visibilidade da imagem: %s com confian√ßa %f" % (image_path, confidence))
        for attempt in range(max_retries):
            button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if button_location is not None:
                x, y = pyautogui.center(button_location)
                logging.info("Imagem vis√≠vel nas coordenadas: x=%d, y=%d" % (x, y))
                return button_location
            else:
                logging.error("Tentativa %d/%d: Imagem n√£o encontrada na tela." % (attempt + 1, max_retries))
                time.sleep(2)
        return None
    except Exception as e:
        logging.error("Erro ao verificar visibilidade da imagem: %s" % str(e))
        return None

def find_and_click_button(image_path, confidence=0.95):
    try:
        while True:
            button_location = verify_image_visibility(image_path, confidence=confidence)
            if button_location is not None:
                x, y = pyautogui.center(button_location)
                logging.info("Bot√£o encontrado nas coordenadas: x=%d, y=%d" % (x, y))
                pyautogui.click(x, y)
                break
            else:
                logging.info("Tentando localizar a imagem novamente...")
                time.sleep(2)
    except Exception as e:
        logging.error("Erro ao localizar o bot√£o: %s" % str(e))

def execute_vsloader(shopping, tipo):
    try:
        # Determina o faturamento_mode baseado no par√¢metro 'tipo'
        if "at√≠p" in tipo.lower():
            faturamento_mode = "AT√çPICOS"
        elif "post" in tipo.lower():
            faturamento_mode = "POSTECIPADO"
        else:
            faturamento_mode = "ANTECIPADO"
        logging.info(f"Iniciando processo para {tipo} (Modo: {faturamento_mode}) no shopping {shopping}")
        # Primeiro gera os arquivos .txt e obt√©m o path de sa√≠da
        output_dir, file_count = generate_txts_from_xls(shopping, tipo)
        logging.info(f"{file_count} arquivos .txt gerados em {output_dir}") 


        # Agora sim, depois de gerar os arquivos, liste-os:
        # Agora sim, depois de gerar os arquivos, liste-os, removendo duplicados:
        all_txt_files = [
            f for f in os.listdir(output_dir)
            if f.lower().endswith('.txt')
        ]

        # Remove duplicidades e ordena a lista
        txt_files = sorted(set(all_txt_files))

        # Atualiza file_count para refletir o n√∫mero real de arquivos √∫nicos
        file_count = len(txt_files)
        logging.info(f"{file_count} arquivos .txt gerados (sem duplicatas) em {output_dir}")

                

        # Determina o variant dinamicamente com base no nome do shopping
        variant = determine_variant(shopping)

        lista_arquivos = [
            "AguaConsumoQSQ",
            "AguaQSQ",
            "ArCondQSQ",
            "CrachaQSQ",
            "DeditizacaoQSQ",
            "DESC_FPPECQSQ",
            "Disp_SistemaQSQ",
            "E_C_QSQ",
            "EnergiaConsumoQSQ",
            "EnergiaQSQ",
            "FPPECQSQ",
            "IPTU_PARC_QSQ",
            "IPTUCotaUnicaQSQ",
            "OutrosQSQ"
        ]

        folder = folder_map.get(shopping, r"C:\Program Files\Victor & Schellenberger_FAT\VSSC_MONTSERRAT")
        up_times = shopping_num_map.get(shopping, 1)

        

        user32 = ctypes.windll.user32
        def get_foreground_window():
            return user32.GetForegroundWindow()

        def wait_for_stable_focus(prev_handle, max_wait=15):
            start_time = time.time()
            while True:
                if time.time() - start_time > max_wait:
                    logging.info("Tempo m√°ximo de espera por foco est√°vel atingido.")
                    break
                time.sleep(1)
                current_handle = get_foreground_window()
                if current_handle != prev_handle:
                    time.sleep(1.5)
                    second_check = get_foreground_window()
                    if second_check != current_handle:
                        prev_handle = second_check
                        continue
                    break

        def wait_for_focus_change(prev_handle, max_wait=40):
            start_time = time.time()
            while True:
                if time.time() - start_time > max_wait:
                    logging.info("Tempo m√°ximo de espera por mudan√ßa de foco atingido.")
                    break
                time.sleep(1)
                current_handle = get_foreground_window()
                if current_handle != prev_handle:
                    break
        
        screen_width, screen_height = pyautogui.size()
        center_x = screen_width // 2
        center_y = screen_height // 2
        
        

            
    


        def importar_encargos(file_count, txt_files, output_dir, folder):
            time.sleep(4)
            # Deixa fixo:
            # - Postecipado nas colunas 1..4
            # - Antecipado nas colunas 6..9
            # - At√≠pico nas colunas 11..14

            # Remove duplicadas e ordena a lista de arquivos TXT
            txt_files = sorted(set(txt_files))
            file_count = len(txt_files)
            logging.info(f"Ap√≥s remo√ß√£o de duplicatas, {file_count} arquivos √∫nicos ser√£o processados.")

            # Divide encargos comuns e fundos FPP
            fpp_files = [f for f in txt_files if "FPP" in f]
            encargos_files = [f for f in txt_files if "FPP" not in f]

            # Inicializa o arquivo de c√°lculos (planilha Excel) dentro da pasta de logs do shopping
            calc_filename = os.path.join(logs_dir, f"{os.path.basename(folder).replace('_HOM','')}_calculos.xlsx")

            # Mapeia as colunas de cada tipo de faturamento
            # POSTECIPADO: colunas 1..4
            # ANTECIPADO: colunas 6..9
            # AT√çPICO: colunas 11..14
            if faturamento_mode.lower() == "postecipado":
                col_tipo = 1
                col_arquivo = 2
                col_encargo = 3
                col_status = 4
            elif faturamento_mode.lower() == "antecipado":
                col_tipo = 6
                col_arquivo = 7
                col_encargo = 8
                col_status = 9
            elif faturamento_mode.lower() == "at√≠picos":
                col_tipo = 11
                col_arquivo = 12
                col_encargo = 13
                col_status = 14

            if not os.path.exists(calc_filename):
                wb_calc = Workbook()
                ws_calc = wb_calc.active
                ws_calc.title = "Encargos"

                # Cabe√ßalhos para Postecipado (colunas 1..4)
                ws_calc.cell(row=1, column=1, value="Tipo de Faturamento")
                ws_calc.cell(row=1, column=2, value="Nome do Arquivo")
                ws_calc.cell(row=1, column=3, value="Encargo")
                ws_calc.cell(row=1, column=4, value="Processamento")

                # Coluna 5 em branco (espa√ßamento)

                # Cabe√ßalhos para Antecipado (colunas 6..9)
                ws_calc.cell(row=1, column=6, value="Tipo de Faturamento")
                ws_calc.cell(row=1, column=7, value="Nome do Arquivo")
                ws_calc.cell(row=1, column=8, value="Encargo")
                ws_calc.cell(row=1, column=9, value="Processamento")

                # Coluna 10 em branco (espa√ßamento)

                # Cabe√ßalhos para At√≠pico (colunas 11..14)
                ws_calc.cell(row=1, column=11, value="Tipo de Faturamento")
                ws_calc.cell(row=1, column=12, value="Nome do Arquivo")
                ws_calc.cell(row=1, column=13, value="Encargo")
                ws_calc.cell(row=1, column=14, value="Processamento")

                wb_calc.save(calc_filename)
            else:
                wb_calc = load_workbook(calc_filename)
                ws_calc = wb_calc.active
                # Limpa somente a coluna de status do tipo atual
                max_row = ws_calc.max_row
                for row in range(2, max_row + 1):
                    ws_calc.cell(row=row, column=col_status, value="")
                wb_calc.save(calc_filename)

            wb_calc = load_workbook(calc_filename)
            ws_calc = wb_calc.active

            # Pr√©-registra na planilha cada encargo como PENDENTE na coluna de Processamento
            for idx in range(1, file_count + 1):
                arquivo_txt = txt_files[idx - 1]
                row_num = idx + 1
                ws_calc.cell(row=row_num, column=col_tipo, value=faturamento_mode)
                ws_calc.cell(row=row_num, column=col_arquivo, value=arquivo_txt)
                ws_calc.cell(row=row_num, column=col_encargo, value=f"Encargo {idx}")
                ws_calc.cell(row=row_num, column=col_status, value="PENDENTE")
            wb_calc.save(calc_filename)

            # Calcula posi√ß√µes nos arquivos (1-based)
            positions_encargos = [i + 1 for i, f in enumerate(txt_files) if f not in fpp_files]
            positions_funds = [i + 1 for i, f in enumerate(txt_files) if f in fpp_files]

            # Loop para processar todos os encargos (exceto FPP)
            while True:
                screen_width, screen_height = pyautogui.size()
                center_x = screen_width // 2
                center_y = screen_height // 2

                pyautogui.moveTo(center_x, center_y)
                pyautogui.click()

                logging.info("Iniciando processo de importa√ß√£o de encargos.")

                # Entra na tela de importa√ß√£o de encargos
                pyautogui.hotkey('alt', 's')
                for _ in range(12):
                    pyautogui.press('right')
                time.sleep(1)

                pyautogui.press('down')
                pyautogui.press('enter')
                time.sleep(0.5)
                pyautogui.press('enter')

                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle)
                wait_for_stable_focus(prev_handle)

                logging.info(f"Iniciando loop principal de importa√ß√£o para {len(positions_encargos)} arquivos.")

                wb_calc = load_workbook(calc_filename)
                ws_calc = wb_calc.active


                base_dir   = os.path.dirname(output_dir)                       # remove ‚ÄúArquivos Cargas‚Äù
                result_dir = os.path.join(base_dir, "Resultado das Cargas")

                if os.path.exists(result_dir):
                    # elimina tudo dentro dela
                    for nome in os.listdir(result_dir):
                        caminho = os.path.join(result_dir, nome)
                        if os.path.isfile(caminho):
                            os.remove(caminho)
                        else:
                            shutil.rmtree(caminho)
                else:
                    os.makedirs(result_dir, exist_ok=True)
                    

                for pos in positions_encargos:
                    time.sleep(2)
                    logging.info(f"Importando encargo de n√∫mero {pos}")
                    arquivo_txt = txt_files[pos - 1]
                    filepath = os.path.join(output_dir, arquivo_txt)

                    # Abre janela "Abrir" via Alt+Space
                    pyautogui.hotkey('alt', 'space')
                    time.sleep(0.3)
                    pyautogui.press('down')
                    time.sleep(0.3)
                    pyautogui.press('enter')
                    pyautogui.moveRel(148, 52)
                    pyautogui.click()
                    pyautogui.click()

                    time.sleep(2)

                    pyperclip.copy(output_dir)
                    pyautogui.hotkey('ctrl', 'v')
                    time.sleep(1)

                    pyautogui.typewrite("\\")
                    time.sleep(2)

                    # Pressiona "down" pos vezes
                    for _ in range(pos):
                        pyautogui.press('down')
                    time.sleep(2)

                    pyautogui.press('enter')

                    while True:
                        screenshot1 = capture_screenshot()
                        extracted1 = analyze_screenshot(screenshot1) or ""
                        time.sleep(2)
                        screenshot2 = capture_screenshot()
                        extracted2 = analyze_screenshot(screenshot2) or ""
                        if extracted1 == "" and extracted2 == "":
                            time.sleep(3)
                            continue
                        if (
                            ("n√£o h√° nenhum modal" in extracted1.lower() or "lobby" in extracted1.lower() or "modal n√£o detectado" in extracted1.lower())
                            and ("n√£o h√° nenhum modal" in extracted2.lower() or "lobby" in extracted2.lower() or "modal n√£o detectado" in extracted2.lower())
                        ):
                            logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints.")
                            time.sleep(3)
                            break
                        combined = extracted1 + " " + extracted2
                        if fuzzy_contains(combined, "Alerta VSSC"):
                            logging.info(f"execute_vsloader: Alerta VSSC. Texto: {combined}")
                            pyautogui.moveTo(center_x, center_y)
                            pyautogui.click()
                            pyautogui.press("esc")
                        else:
                            break

                    pyautogui.hotkey('alt', 'space')
                    time.sleep(0.3)
                    pyautogui.press('down')
                    time.sleep(0.3)
                    pyautogui.press('enter')
                    pyautogui.moveRel(5, 100)
                    pyautogui.click()
                    pyautogui.click()

                    time.sleep(3)
                    for _ in range(2):
                        pyautogui.press('tab')
                    time.sleep(0.5)
                    pyautogui.press("enter")
                    prev_handle = get_foreground_window()
                    wait_for_focus_change(prev_handle)
                    wait_for_stable_focus(prev_handle)


                    while True:
                        screenshot1 = capture_screenshot()
                        extracted1 = analyze_screenshot(screenshot1) or ""
                        time.sleep(2)
                        screenshot2 = capture_screenshot()
                        extracted2 = analyze_screenshot(screenshot2) or ""
                        if extracted1 == "" and extracted2 == "":
                            time.sleep(3)
                            continue
                        if (
                            ("n√£o h√° nenhum modal" in extracted1.lower() or "lobby" in extracted1.lower() or "modal n√£o detectado" in extracted1.lower())
                            and ("n√£o h√° nenhum modal" in extracted2.lower() or "lobby" in extracted2.lower() or "modal n√£o detectado" in extracted2.lower())
                        ):
                            logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints.")
                            time.sleep(2)
                            break
                        combined = extracted1 + " " + extracted2
                        if fuzzy_contains(combined, "Alerta VSSC"):
                            logging.info(f"execute_vsloader: A linha ser√° desprezada. Texto: {combined}")
                            pyautogui.moveTo(center_x, center_y)
                            pyautogui.click()
                            pyautogui.press("esc")

                            
                            
                            # monta o resultado em ‚Äú‚Ä¶\AT√çPICAS\Resultado das Cargas‚Äù
                            
                            prev_handle = get_foreground_window()
                            wait_for_focus_change(prev_handle)
                            wait_for_stable_focus(prev_handle)
                            base_dir   = os.path.dirname(output_dir)                       # remove ‚ÄúArquivos Cargas‚Äù
                            result_dir = os.path.join(base_dir, "Resultado das Cargas")
                            os.makedirs(result_dir, exist_ok=True)

                            # captura e move o screenshot para l√°
                            screenshot_path = capture_screenshot()
                            dest_path       = os.path.join(result_dir, os.path.basename(screenshot_path))
                            shutil.move(screenshot_path, dest_path)

                            logging.info(f"Screenshot de 'A linha ser√° desprezada' salva em: {dest_path}")
                            time.sleep(1)
                            for _ in range(2):
                                pyautogui.press('tab')
                                time.sleep(0.5)
                            pyautogui.press("enter")
                            time.sleep(1)
                            break
                            
                        else:
                            logging.info(f"execute_vsloader: N√£o achou o alerta de linha ser√° desprezada. Texto: {combined}")
                            time.sleep(10)
                            base_dir   = os.path.dirname(output_dir)                       # remove ‚ÄúArquivos Cargas‚Äù
                            result_dir = os.path.join(base_dir, "Resultado das Cargas")
                            os.makedirs(result_dir, exist_ok=True)

                            # captura e move o screenshot para l√°
                            screenshot_path = capture_screenshot()
                            dest_path       = os.path.join(result_dir, os.path.basename(screenshot_path))
                            shutil.move(screenshot_path, dest_path)

                            logging.info(f"Screenshot de 'A linha ser√° desprezada' salva em: {dest_path}")
                            time.sleep(1)
                            for _ in range(2):
                                pyautogui.press('tab')
                                time.sleep(0.5)
                            pyautogui.press("enter")
                            time.sleep(1)
                            break
                          

                    

                    while True:
                        screenshot1 = capture_screenshot()
                        extracted1 = analyze_screenshot(screenshot1) or ""
                        time.sleep(2)
                        screenshot2 = capture_screenshot()
                        extracted2 = analyze_screenshot(screenshot2) or ""
                        if extracted1 == "" and extracted2 == "":
                            time.sleep(3)
                            continue
                        if (
                            ("n√£o h√° nenhum modal" in extracted1.lower() or "lobby" in extracted1.lower() or "modal n√£o detectado" in extracted1.lower())
                            and ("n√£o h√° nenhum modal" in extracted2.lower() or "lobby" in extracted2.lower() or "modal n√£o detectado" in extracted2.lower())
                        ):
                            logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints.")
                            time.sleep(2)
                            break
                        combined = extracted1 + " " + extracted2
                        if fuzzy_contains(combined, "Alerta VSSC") and fuzzy_contains(combined, "Confirma a atualiza√ß√£o"):
                            logging.info(f"execute_vsloader: Alerta VSSC. Texto: {combined}")
                            pyautogui.moveTo(center_x, center_y)
                            pyautogui.click()
                            pyautogui.press("enter")
                            ws_calc.cell(row=pos + 1, column=col_status, value="Encargo importado")
                            wb_calc.save(calc_filename)
                            time.sleep(0.5)
                            prev_handle = get_foreground_window()
                            wait_for_focus_change(prev_handle)
                            wait_for_stable_focus(prev_handle)
                            pyautogui.moveTo(center_x, center_y)
                            pyautogui.click()
                            pyautogui.press("esc")
                            time.sleep(3)

                        elif fuzzy_contains(combined, "Alerta VSSC"):
                            logging.info(f"execute_vsloader: Alerta VSSC. Texto: {combined}")
                            pyautogui.moveTo(center_x, center_y)
                            pyautogui.click()
                            pyautogui.press("esc")
                       
                        elif fuzzy_contains(combined, "IMPORTAR VALORES DE ENCARGOS"):
                            logging.info(f"execute_vsloader: Importar valores de encargos. Texto: {combined}")
                            
                            break
                        elif (
                            fuzzy_contains(combined, "Compet√™ncia de trabalho:")
                            and fuzzy_contains(combined, "Per√≠odo Fechado")
                            and fuzzy_contains(combined, "(Faturamento)")
                        ):
                            logging.info(f"execute_vsloader: Lobby identificado. Texto: {combined}")
                            break
                        else:
                            logging.info(f"execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto: {combined}")

                # Fecha modal de encargos com 5 ESCs
                for _ in range(5):
                    pyautogui.press('esc')
                    time.sleep(0.5)
                break

            # Importa√ß√£o de fundos FPP
            for pos in positions_funds:
                time.sleep(2)
                logging.info(f"Importando fundo FPP de n√∫mero {pos}")

                # Entra na tela de importa√ß√£o de fundos (FPP)
                pyautogui.hotkey('alt', 's')
                for _ in range(12):
                    pyautogui.press('right')
                time.sleep(1)

                pyautogui.press('down')
                pyautogui.press('enter')
                time.sleep(0.5)
                for _ in range(2):
                    pyautogui.press('down')
                pyautogui.press('enter')

                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle)
                wait_for_stable_focus(prev_handle)

                wb_calc = load_workbook(calc_filename)
                ws_calc = wb_calc.active

                # Mesma l√≥gica de importa√ß√£o de encargos
                for _ in range(positions_funds.count(pos)):
                    arquivo_txt = txt_files[pos - 1]
                    filepath = os.path.join(output_dir, arquivo_txt)

                    # Abre janela "Abrir" via Alt+Space
                    pyautogui.hotkey('alt', 'space')
                    time.sleep(0.3)
                    pyautogui.press('down')
                    time.sleep(0.3)
                    pyautogui.press('enter')
                    pyautogui.moveRel(148, 52)
                    pyautogui.click()
                    pyautogui.click()

                    time.sleep(2)

                    pyperclip.copy(output_dir)
                    pyautogui.hotkey('ctrl', 'v')
                    time.sleep(1)

                    pyautogui.typewrite("\\")
                    time.sleep(2)

                    for _ in range(pos):
                        pyautogui.press('down')
                    time.sleep(2)

                    pyautogui.press('enter')

                    # Reuso dos mesmos loops de captura e a√ß√µes
                    while True:
                        screenshot1 = capture_screenshot()
                        extracted1 = analyze_screenshot(screenshot1) or ""
                        time.sleep(2)
                        screenshot2 = capture_screenshot()
                        extracted2 = analyze_screenshot(screenshot2) or ""
                        if extracted1 == "" and extracted2 == "":
                            time.sleep(3)
                            continue
                        if (
                            ("n√£o h√° nenhum modal" in extracted1.lower() or "lobby" in extracted1.lower() or "modal n√£o detectado" in extracted1.lower())
                            and ("n√£o h√° nenhum modal" in extracted2.lower() or "lobby" in extracted2.lower() or "modal n√£o detectado" in extracted2.lower())
                        ):
                            logging.info("execute_vsloader: Lobby identificado em FPP.")
                            time.sleep(3)
                            break
                        combined = extracted1 + " " + extracted2
                        if fuzzy_contains(combined, "Alerta VSSC"):
                            pyautogui.moveTo(center_x, center_y)
                            pyautogui.click()
                            pyautogui.press("esc")
                        else:
                            break

                    for _ in range(5):
                        pyautogui.press('esc')
                        time.sleep(0.5)













        # =============================
        # A partir daqui, faz os c√°lculos sequenciais
        # =============================
        def executar_calculos():
            # while True:
            time.sleep(3)
            # Inicializa vari√°veis de ambiente necess√°rias
            screen_width, screen_height = pyautogui.size()
            center_x = screen_width // 2
            center_y = screen_height // 2
            for _ in range(3):
                pyautogui.press('esc')
                time.sleep(0.3)

            # Inicializa o arquivo de c√°lculos (planilha Excel)
            calc_filename = os.path.join(logs_dir, f"{os.path.basename(folder).replace('_HOM','')}_calculos.xlsx")
            # Define as colunas de acordo com o tipo de faturamento
            if faturamento_mode.lower() == "postecipado":
                col_tipo, col_arquivo, col_encargo, col_status = 1, 2, 3, 4
            elif faturamento_mode.lower() == "antecipado":
                col_tipo, col_arquivo, col_encargo, col_status = 6, 7, 8, 9
            elif faturamento_mode.lower() == "at√≠picos":
                col_tipo, col_arquivo, col_encargo, col_status = 11, 12, 13, 14
            else:
                logging.error("Faturamento_mode n√£o reconhecido!")
                return

            if not os.path.exists(calc_filename):
                wb_calc = Workbook()
                ws_calc = wb_calc.active
                ws_calc.title = "Encargos"
                # Cabe√ßalhos para Postecipado (colunas 1..4)
                ws_calc.cell(row=1, column=1, value="Tipo de Faturamento")
                ws_calc.cell(row=1, column=2, value="Nome do Arquivo")
                ws_calc.cell(row=1, column=3, value="Encargo")
                ws_calc.cell(row=1, column=4, value="Processamento")
                # Cabe√ßalhos para Antecipado (colunas 6..9)
                ws_calc.cell(row=1, column=6, value="Tipo de Faturamento")
                ws_calc.cell(row=1, column=7, value="Nome do Arquivo")
                ws_calc.cell(row=1, column=8, value="Encargo")
                ws_calc.cell(row=1, column=9, value="Processamento")
                # Cabe√ßalhos para At√≠pico (colunas 11..14)
                ws_calc.cell(row=1, column=11, value="Tipo de Faturamento")
                ws_calc.cell(row=1, column=12, value="Nome do Arquivo")
                ws_calc.cell(row=1, column=13, value="Encargo")
                ws_calc.cell(row=1, column=14, value="Processamento")
                wb_calc.save(calc_filename)
            else:
                wb_calc = load_workbook(calc_filename)

            if "Contas x Fases" not in wb_calc.sheetnames:
                ws_contas = wb_calc.create_sheet("Contas x Fases")
                ws_contas.cell(row=1, column=1, value="Tipo de Faturamento")
                ws_contas.cell(row=1, column=2, value="Conta")
                ws_contas.cell(row=1, column=3, value="Fase")
                ws_contas.cell(row=1, column=4, value="Status")
                wb_calc.save(calc_filename)
            else:
                ws_contas = wb_calc["Contas x Fases"]

            # Preenche automaticamente as combina√ß√µes (conta, fase) na aba "Contas x Fases"
            combos = []
            if faturamento_mode.upper() == "POSTECIPADO":
                combos = [
                    ("310100", get_phase(shopping, "310100", faturamento_mode, variant)),
                    ("210100", get_phase(shopping, "210100", faturamento_mode, variant)),
                    ("200101", get_phase(shopping, "200101", faturamento_mode, variant)),
                    ("200106" if faturamento_mode.upper() == "POSTECIPADO" else "200133",
                    get_phase(shopping, "200106", faturamento_mode, variant) if faturamento_mode.upper() == "POSTECIPADO"
                    else get_phase(shopping, "200133", faturamento_mode, variant)),
                    ("410900", get_phase(shopping, "410900", faturamento_mode, variant)),
                    ("411000", get_phase(shopping, "411000", faturamento_mode, variant))
                ]
                # Adiciona fases extras para Postecipado
                extra_calcs = get_extra_phases(shopping, faturamento_mode, variant)
                for extra in extra_calcs:
                    combos.append(extra)
            elif faturamento_mode.upper() == "ANTECIPADO":
                combos = [
                    ("310100", get_phase(shopping, "310100", faturamento_mode, variant)),
                    ("210100", get_phase(shopping, "210100", faturamento_mode, variant)),
                    ("200101", get_phase(shopping, "200101", faturamento_mode, variant)),
                    ("200133", get_phase(shopping, "200133", faturamento_mode, variant)),
                    ("410900", get_phase(shopping, "410900", faturamento_mode, variant)),
                    ("411000", get_phase(shopping, "411000", faturamento_mode, variant))
                ]
                # Adiciona fases extras para Antecipado
                extra_calcs = get_extra_phases(shopping, faturamento_mode, variant)
                for extra in extra_calcs:
                    combos.append(extra)
            elif faturamento_mode.upper() == "AT√çPICOS":
                combos = [
                    ("311101", get_phase(shopping, "311101", faturamento_mode, variant)),
                    ("211101", get_phase(shopping, "211101", faturamento_mode, variant)),
                    ("300101", get_phase(shopping, "300101", faturamento_mode, variant)),
                    ("300101_mall", get_phase(shopping, "300101_mall", faturamento_mode, variant)),
                    ("311128", get_phase(shopping, "311128", faturamento_mode, variant)),
                    ("311129", get_phase(shopping, "311129", faturamento_mode, variant))
                ]
                # Adiciona fases extras para At√≠picos
                extra_calcs = get_extra_phases(shopping, faturamento_mode, variant)
                for extra in extra_calcs:
                    combos.append(extra)

            existing_set = set()
            max_row_contas = ws_contas.max_row
            for rowi in range(2, max_row_contas + 1):
                tft = ws_contas.cell(row=rowi, column=1).value
                cta = ws_contas.cell(row=rowi, column=2).value
                fse = ws_contas.cell(row=rowi, column=3).value
                existing_set.add((tft, cta, fse))
            row_pointer = max_row_contas + 1
            for (conta, fase) in combos:
                if fase is None:
                    continue
                if (faturamento_mode.upper(), conta, fase) not in existing_set:
                    ws_contas.cell(row=row_pointer, column=1, value=faturamento_mode.upper())
                    ws_contas.cell(row=row_pointer, column=2, value=conta)
                    ws_contas.cell(row=row_pointer, column=3, value=fase)
                    ws_contas.cell(row=row_pointer, column=4, value="PENDENTE")
                    row_pointer += 1
            wb_calc.save(calc_filename)

            if faturamento_mode.upper() == "AT√çPICOS":
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 311101")
                # Sequ√™ncia espec√≠fica para contas at√≠picas, na ordem solicitada:
                # 311101 (Encargo Comum At√≠pico),
                # 211101 (FPP At√≠pico),
                # 300101 (Aluguel Complementar At√≠pico),
                # 300101 MALL (Aluguel Complementar MALL),
                # 311128 (Aux Taxa ADM At√≠pico),
                # 311129 (Taxa ADM At√≠pico)

                # ================ BLOCO 1: CALCULAR ENCARGO COMUM (CONTA 310100) ================
                time.sleep(2)
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=10)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("311101")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(12)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_fpp = get_phase(shopping, "311101", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_fpp)
                logging.info(f"Clicando na fase {phase_fpp}")
                time.sleep(5)
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)

                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    
                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3) 
                        
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        
                        # Marca como "C√°lculo efetuado" no "Contas x Fases"
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                        
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")  
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia do trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Quota Ordin√°ria"):
                        logging.info("execute_vsloader: Calculas Valores. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        break
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # # ================ BLOCO 2: CALCULAR FPP (CONTA 210100) ================
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 211101")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(3):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("211101")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_fpp = get_phase(shopping, "211101", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_fpp)
                logging.info(f"Clicando na fase {phase_fpp}")
                time.sleep(5)
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)

                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Quota Ordin√°ria"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        break
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # ================ BLOCO 3: CALCULAR ALUGUEL M√çNIMO (CONTA 200101) ================
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 300101")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(1):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                for _ in range(1):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("300101")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_desconto = get_phase(shopping, "300101", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_desconto)
                logging.info(f"Clicando na fase {phase_desconto}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Alerta VSSC Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Desconto Pontualidade"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        break
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # # ================ BLOCO 4: CALCULAR DESCONTO TERMO ACORDADO (CONTA 200106) ================
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 300101")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(1):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                for _ in range(1):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("300101")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_desconto = get_phase(shopping, "300101_mall", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_desconto)
                logging.info(f"Clicando na fase {phase_desconto}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Desconto Pontualidade"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        break
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # # ========== BLOCO 5: CALCULAR AUXILIAR DA TAXA ADM (CONTA 410900) ==========
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 311128")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("311128")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_aux = get_phase(shopping, "311128", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_aux)
                logging.info(f"Clicando na fase {phase_aux}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Taxa Administrativa"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # # ========== BLOCO 6: CALCULAR TAXA ADM (CONTA 411000) ==========
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 311129")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("311129")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_adm = get_phase(shopping, "311129", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_adm)
                logging.info(f"Clicando na fase {phase_adm}")
                time.sleep(5)
            
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                    
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Tava Administrativa"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        break
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

    ##################### FIM DOS BLOCOS ATIPICOS #######################

                time.sleep(2)
                pyautogui.hotkey('alt', 's')
                for _ in range(9):
                    pyautogui.press('right')
                    time.sleep(0.3)
                
                pyautogui.press('enter')
                time.sleep(2)
                prev_handle = get_foreground_window()
                wait_for_stable_focus(prev_handle)
                pyautogui.moveTo(center_x, center_y)
                pyautogui.click()
                time.sleep(0.3)
                pyautogui.press('enter')
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A op√ß√£o que associa"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ENTER. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ENTER. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                    
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ENTER. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Tava Administrativa"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "Tecle ESC para sair") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        break
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)
                
                

                
            else:
            # ================ BLOCO 1: CALCULAR ENCARGO COMUM (CONTA 310100) ================
                logging.info(f"Iniciando c√°lculo dos encargos - conta 310100")
                time.sleep(2)
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                time.sleep(2)
                prev_handle = get_foreground_window()
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("310100")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_encargo = get_phase(shopping, "310100", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_encargo)  # Fase para Encargo Comum
                logging.info(f"Clicando na fase {phase_encargo}")
                time.sleep(5)
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)

                

                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2

                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                                
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                    
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Encargos Comuns"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # ================ BLOCO 2: CALCULAR FPP (CONTA 210100) ================
                logging.info(f"Iniciando c√°lculo dos encargos at√≠picos - conta 210100")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(3):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("210100")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_fpp = get_phase(shopping, "210100", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_fpp)
                logging.info(f"Clicando na fase {phase_fpp}")
                time.sleep(5)
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)

                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                    
                        
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Quota Ordin√°ria"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # ================ BLOCO 3: CALCULAR ALUGUEL M√çNIMO (CONTA 200101) ================
                logging.info(f"Iniciando c√°lculo dos encargos - conta 200101")
                time.sleep(2)
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(1):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("200101")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_aluguel = get_phase(shopping, "200101", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_aluguel)
                logging.info(f"Clicando na fase {phase_aluguel}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")

                       
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Aluguel Minimo"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # ================ BLOCO 4: CALCULAR DESCONTO TERMO ACORDADO (CONTA 200106) ================
                logging.info(f"Iniciando c√°lculo dos encargos - conta 200106")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(1):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                if faturamento_mode.upper() == "POSTECIPADO":
                    pyautogui.typewrite("200106")
                elif faturamento_mode.upper() == "ANTECIPADO":
                    pyautogui.typewrite("200133")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                if faturamento_mode.upper() == "POSTECIPADO":
                    phase_desconto = get_phase(shopping, "200106", faturamento_mode, variant)
                elif faturamento_mode.upper() == "ANTECIPADO":
                    phase_desconto = get_phase(shopping, "200133", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_desconto)
                logging.info(f"Clicando na fase {phase_desconto}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2

                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        
                    

                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Desconto Pontualidade"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)
                # ========== BLOCO 5: CALCULAR AUXILIAR DA TAXA ADM (CONTA 410900) ==========
                logging.info(f"Iniciando c√°lculo dos encargos - conta 410900")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("410900")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_aux = get_phase(shopping, "410900", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_aux)
                logging.info(f"Clicando na fase {phase_aux}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")

                        
                        
                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Taxa Administrativa"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # ========== BLOCO 6: CALCULAR TAXA ADM (CONTA 411000) ==========
                logging.info(f"Iniciando c√°lculo dos encargos - conta 411000")
                pyautogui.hotkey('alt', 's')
                for _ in range(6):
                    pyautogui.press('right')
                    time.sleep(0.3)
                for _ in range(2):
                    pyautogui.press('down')
                    time.sleep(0.3)
                pyautogui.press('right')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                prev_handle = get_foreground_window()
                wait_for_focus_change(prev_handle, max_wait=30)
                wait_for_stable_focus(prev_handle)
                pyautogui.hotkey('ctrl', 'f')
                time.sleep(0.5)
                pyautogui.typewrite("411000")
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(-161, 33)
                pyautogui.click()
                pyautogui.click()
                time.sleep(10)
                pyautogui.hotkey('alt', 'space')
                time.sleep(0.3)
                pyautogui.press('down')
                time.sleep(0.3)
                pyautogui.press('enter')
                pyautogui.moveRel(95, 37)
                pyautogui.click()
                pyautogui.click()
                time.sleep(0.5)
                phase_aux = get_phase(shopping, "411000", faturamento_mode, variant)
                click_fase_tipo1(shopping, phase_aux)
                logging.info(f"Clicando na fase {phase_aux}")
                time.sleep(5)
                
                for _ in range(4):
                    pyautogui.press('tab')
                pyautogui.press('enter')
                time.sleep(2)
                while True:
                    screenshot1 = capture_screenshot()
                    extracted1 = analyze_screenshot(screenshot1)
                    time.sleep(2)
                    screenshot2 = capture_screenshot()
                    extracted2 = analyze_screenshot(screenshot2)

                    # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                    if extracted1 is None or not isinstance(extracted1, str):
                        extracted1 = ""
                    if extracted2 is None or not isinstance(extracted2, str):
                        extracted2 = ""

                    if not extracted1 and not extracted2:
                        time.sleep(3)
                        continue

                    if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                        and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                        logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                        time.sleep(3)
                        break
                    combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                    screen_width, screen_height = pyautogui.size()
                    center_x = screen_width // 2
                    center_y = screen_height // 2
                    if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                        time.sleep(1)
                        for _ in range(4):
                            pyautogui.press("esc")
                            time.sleep(0.3)
                        logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                        wb_calc = load_workbook(calc_filename)
                        if "Contas x Fases" not in wb_calc.sheetnames:
                            ws_contas = wb_calc.create_sheet("Contas x Fases")
                        else:
                            ws_contas = wb_calc["Contas x Fases"]
                        max_r = ws_contas.max_row
                        for r in range(2, max_r+1):
                            tfat = ws_contas.cell(row=r, column=1).value
                            cta = ws_contas.cell(row=r, column=2).value
                            fse = ws_contas.cell(row=r, column=3).value
                            if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                break
                        wb_calc.save(calc_filename)
                        continue
                    elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                        logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        
                    

                    elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                        logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("enter")
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                        logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        for _ in range(3):
                            pyautogui.press('enter')
                            time.sleep(0.3)

                    elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Taxa Administrativa"):
                        logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                        break
                        
                    elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                        logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                        pyautogui.moveTo(center_x, center_y)
                        pyautogui.click()
                        pyautogui.press("esc")
                        
                    elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                        logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                        logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                        fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                        fuzzy_contains(combined_extracted, "(Faturamento)")):
                        logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                        time.sleep(3)
                        break
                    else:
                        logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                for _ in range(3):
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press('esc')
                    time.sleep(0.3)

                # ========================= Execu√ß√£o dos C√°lculos Extras =========================
                logging.info(f"Iniciando c√°lculo dos encargos extras")
                extra_calcs = get_extra_phases(shopping, faturamento_mode, variant)
                if extra_calcs:
                    for account, extra_phase in extra_calcs:
                        if account == "200101":
                            logging.info(f"Iniciando c√°lculo dos encargos extras - conta 200101")
                            # Replicar o bloco de "Calcular Aluguel M√≠nimo"
                            time.sleep(2)
                            pyautogui.hotkey('alt', 's')
                            for _ in range(6):
                                pyautogui.press('right')
                                time.sleep(0.3)
                            for _ in range(1):
                                pyautogui.press('down')
                                time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            prev_handle = get_foreground_window()
                            wait_for_focus_change(prev_handle, max_wait=30)
                            wait_for_stable_focus(prev_handle)
                            pyautogui.hotkey('ctrl', 'f')
                            time.sleep(0.5)
                            pyautogui.typewrite(account)
                            pyautogui.press('enter')
                            time.sleep(1)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(-161, 33)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(10)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(95, 37)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(0.5)
                            click_fase_tipo1(shopping, extra_phase)
                            logging.info(f"Clicando na fase {extra_phase}")
                            time.sleep(5)
                            for _ in range(4):
                                pyautogui.press('tab')
                            pyautogui.press('enter')
                            time.sleep(2)
                            while True:
                                screenshot1 = capture_screenshot()
                                extracted1 = analyze_screenshot(screenshot1)
                                time.sleep(2)
                                screenshot2 = capture_screenshot()
                                extracted2 = analyze_screenshot(screenshot2)

                                # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                                if extracted1 is None or not isinstance(extracted1, str):
                                    extracted1 = ""
                                if extracted2 is None or not isinstance(extracted2, str):
                                    extracted2 = ""

                                if not extracted1 and not extracted2:
                                    time.sleep(3)
                                    continue

                                if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                                    and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                                    logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                                    time.sleep(3)
                                    break
                                combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                                screen_width, screen_height = pyautogui.size()
                                center_x = screen_width // 2
                                center_y = screen_height // 2
                                if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                                    time.sleep(1)
                                    for _ in range(4):
                                        pyautogui.press("esc")
                                        time.sleep(0.3)
                                    logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                                    wb_calc = load_workbook(calc_filename)
                                    if "Contas x Fases" not in wb_calc.sheetnames:
                                        ws_contas = wb_calc.create_sheet("Contas x Fases")
                                    else:
                                        ws_contas = wb_calc["Contas x Fases"]
                                    max_r = ws_contas.max_row
                                    for r in range(2, max_r+1):
                                        tfat = ws_contas.cell(row=r, column=1).value
                                        cta = ws_contas.cell(row=r, column=2).value
                                        fse = ws_contas.cell(row=r, column=3).value
                                        if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                            ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                            break
                                    wb_calc.save(calc_filename)
                                    continue 
                                elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                                    logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    
                                   
                                elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                                    logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                                    logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    for _ in range(3):
                                        pyautogui.press('enter')
                                        time.sleep(0.3)

                                elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Aluguel Minimo"):
                                    logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                                    break
                                    
                                elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                                    logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("esc")
                                    
                                elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                                    logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                                    logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                                    fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                                    fuzzy_contains(combined_extracted, "(Faturamento)")):
                                    logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                else:
                                    logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                                time.sleep(3)
                            
                            for _ in range(3):
                                pyautogui.moveTo(center_x, center_y)
                                pyautogui.click()
                                pyautogui.press('esc')
                                time.sleep(0.3)

                        elif account == "310100":
                            logging.info(f"Iniciando c√°lculo dos encargos extras - conta 310100")
                            # Replicar o bloco de "Calcular Encargo Comum"
                            screen_width, screen_height = pyautogui.size()
                            center_x = screen_width // 2
                            center_y = screen_height // 2
                            time.sleep(2)
                            pyautogui.hotkey('alt', 's')
                            for _ in range(6):
                                pyautogui.press('right')
                                time.sleep(0.3)
                            for _ in range(2):
                                pyautogui.press('down')
                                time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            time.sleep(2)
                            prev_handle = get_foreground_window()
                            wait_for_stable_focus(prev_handle)
                            pyautogui.hotkey('ctrl', 'f')
                            time.sleep(0.5)
                            pyautogui.typewrite(account)
                            pyautogui.press('enter')
                            time.sleep(1)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(-161, 33)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(10)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(95, 37)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(0.5)
                            click_fase_tipo1(shopping, extra_phase)
                            logging.info(f"Clicando na fase {extra_phase}")
                            time.sleep(5)
                            for _ in range(4):
                                pyautogui.press('tab')
                            pyautogui.press('enter')
                            time.sleep(2)
                            while True:
                                screenshot1 = capture_screenshot()
                                extracted1 = analyze_screenshot(screenshot1)
                                time.sleep(2)
                                screenshot2 = capture_screenshot()
                                extracted2 = analyze_screenshot(screenshot2)

                                # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                                if extracted1 is None or not isinstance(extracted1, str):
                                    extracted1 = ""
                                if extracted2 is None or not isinstance(extracted2, str):
                                    extracted2 = ""

                                if not extracted1 and not extracted2:
                                    time.sleep(3)
                                    continue

                                if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                                    and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                                    logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                                    time.sleep(3)
                                    break
                                combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                                screen_width, screen_height = pyautogui.size()
                                center_x = screen_width // 2
                                center_y = screen_height // 2

                                if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                                    time.sleep(1)
                                    for _ in range(4):
                                        pyautogui.press("esc")
                                        time.sleep(0.3)
                                    logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                                    wb_calc = load_workbook(calc_filename)
                                    if "Contas x Fases" not in wb_calc.sheetnames:
                                        ws_contas = wb_calc.create_sheet("Contas x Fases")
                                    else:
                                        ws_contas = wb_calc["Contas x Fases"]
                                    max_r = ws_contas.max_row
                                    for r in range(2, max_r+1):
                                        tfat = ws_contas.cell(row=r, column=1).value
                                        cta = ws_contas.cell(row=r, column=2).value
                                        fse = ws_contas.cell(row=r, column=3).value
                                        if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                            ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                            break
                                    wb_calc.save(calc_filename)
                                    continue
                                elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                                    logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    
                                     
                                elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                                    logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                                    logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    for _ in range(3):
                                        pyautogui.press('enter')
                                        time.sleep(0.3)

                                elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Encargos Comuns"):
                                    logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                                    break
                                    
                                elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                                    logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("esc")
                                    
                                elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                                    logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                                    logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                                    fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                                    fuzzy_contains(combined_extracted, "(Faturamento)")):
                                    logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                else:
                                    logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                                time.sleep(3)
                            
                            for _ in range(3):
                                pyautogui.moveTo(center_x, center_y)
                                pyautogui.click()
                                pyautogui.press('esc')
                                time.sleep(0.3)

                        elif account == "200119":
                            logging.info(f"Iniciando c√°lculo dos encargos extras - conta 200119")
                            # Espa√ßo Anunciante (Antecipado)
                            time.sleep(2)
                            pyautogui.hotkey('alt', 's')
                            for _ in range(6):
                                pyautogui.press('right')
                                time.sleep(0.3)
                            for _ in range(1):
                                pyautogui.press('down')
                                time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            for _ in range(2):
                                pyautogui.press('down')
                                time.sleep(0.3)
                            pyautogui.press('enter')
                            prev_handle = get_foreground_window()
                            wait_for_focus_change(prev_handle, max_wait=30)
                            wait_for_stable_focus(prev_handle)
                            pyautogui.hotkey('ctrl', 'f')
                            time.sleep(0.5)
                            pyautogui.typewrite(account)
                            pyautogui.press('enter')
                            time.sleep(1)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(-161, 33)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(10)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(95, 37)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(0.5)
                            click_fase_tipo1(shopping, extra_phase)
                            logging.info(f"Clicando na fase {extra_phase}")
                            time.sleep(5)
                            for _ in range(4):
                                pyautogui.press('tab')
                            pyautogui.press('enter')
                            time.sleep(2)
                            while True:
                                screenshot1 = capture_screenshot()
                                extracted1 = analyze_screenshot(screenshot1)
                                time.sleep(2)
                                screenshot2 = capture_screenshot()
                                extracted2 = analyze_screenshot(screenshot2)

                                # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                                if extracted1 is None or not isinstance(extracted1, str):
                                    extracted1 = ""
                                if extracted2 is None or not isinstance(extracted2, str):
                                    extracted2 = ""

                                if not extracted1 and not extracted2:
                                    time.sleep(3)
                                    continue

                                if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                                    and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                                    logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                                    time.sleep(3)
                                    break
                                combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                                screen_width, screen_height = pyautogui.size()
                                center_x = screen_width // 2
                                center_y = screen_height // 2

                                if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                                    time.sleep(1)
                                    for _ in range(4):
                                        pyautogui.press("esc")
                                        time.sleep(0.3)
                                    logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                                    wb_calc = load_workbook(calc_filename)
                                    if "Contas x Fases" not in wb_calc.sheetnames:
                                        ws_contas = wb_calc.create_sheet("Contas x Fases")
                                    else:
                                        ws_contas = wb_calc["Contas x Fases"]
                                    max_r = ws_contas.max_row
                                    for r in range(2, max_r+1):
                                        tfat = ws_contas.cell(row=r, column=1).value
                                        cta = ws_contas.cell(row=r, column=2).value
                                        fse = ws_contas.cell(row=r, column=3).value
                                        if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                            ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                            break
                                    wb_calc.save(calc_filename)
                                    continue
                                elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                                    logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                     
                                    
                                elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                                    logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                                    logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    for _ in range(3):
                                        pyautogui.press('enter')
                                        time.sleep(0.3)

                                elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Espa√ßo para Anunciante"):
                                    logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                                    break
                                    
                                elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                                    logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("esc")
                                    
                                elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                                    logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                                    logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                                    fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                                    fuzzy_contains(combined_extracted, "(Faturamento)")):
                                    logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                else:
                                    logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                                time.sleep(3)
                            
                            for _ in range(3):
                                pyautogui.moveTo(center_x, center_y)
                                pyautogui.click()
                                pyautogui.press('esc')
                                time.sleep(0.3)

                        elif account == "200156":
                            logging.info(f"Iniciando c√°lculo dos encargos extras - conta 200156")
                            # M√≠dia Inaugural (Antecipado)
                            time.sleep(2)
                            pyautogui.hotkey('alt', 's')
                            for _ in range(6):
                                pyautogui.press('right')
                                time.sleep(0.3)
                            for _ in range(1):
                                pyautogui.press('down')
                                time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('right')
                            time.sleep(0.3)
                            for _ in range(2):
                                pyautogui.press('down')
                                time.sleep(0.3)
                            pyautogui.press('enter')
                            prev_handle = get_foreground_window()
                            wait_for_focus_change(prev_handle, max_wait=30)
                            wait_for_stable_focus(prev_handle)
                            pyautogui.hotkey('ctrl', 'f')
                            time.sleep(0.5)
                            pyautogui.typewrite(account)
                            pyautogui.press('enter')
                            time.sleep(1)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(-161, 33)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(10)
                            pyautogui.hotkey('alt', 'space')
                            time.sleep(0.3)
                            pyautogui.press('down')
                            time.sleep(0.3)
                            pyautogui.press('enter')
                            pyautogui.moveRel(95, 37)
                            pyautogui.click()
                            pyautogui.click()
                            time.sleep(0.5)
                            click_fase_tipo1(shopping, extra_phase)
                            logging.info(f"Clicando na fase {extra_phase}")
                            time.sleep(5)
                            for _ in range(4):
                                pyautogui.press('tab')
                            pyautogui.press('enter')
                            time.sleep(2)
                            while True:
                                screenshot1 = capture_screenshot()
                                extracted1 = analyze_screenshot(screenshot1)
                                time.sleep(2)
                                screenshot2 = capture_screenshot()
                                extracted2 = analyze_screenshot(screenshot2)

                                # Garante que extracted1 e extracted2 sejam strings para evitar erro de chamada de .lower()
                                if extracted1 is None or not isinstance(extracted1, str):
                                    extracted1 = ""
                                if extracted2 is None or not isinstance(extracted2, str):
                                    extracted2 = ""

                                if not extracted1 and not extracted2:
                                    time.sleep(3)
                                    continue

                                if ((("n√£o h√° nenhum modal" in extracted1.lower()) or ("lobby" in extracted1.lower()) or ("modal n√£o detectado" in extracted1.lower()))
                                    and (("n√£o h√° nenhum modal" in extracted2.lower()) or ("lobby" in extracted2.lower()) or ("modal n√£o detectado" in extracted2.lower()))):
                                    logging.info("execute_vsloader: Lobby identificado (nenhum modal detectado) em ambos os prints. Texto 1: %s Texto 2: %s", extracted1, extracted2)
                                    time.sleep(3)
                                    break
                                combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                                screen_width, screen_height = pyautogui.size()
                                center_x = screen_width // 2
                                center_y = screen_height // 2

                                if fuzzy_contains(combined_extracted, "Deseja visualizar o arquivo de LOG"):
                                    time.sleep(1)
                                    for _ in range(4):
                                        pyautogui.press("esc")
                                        time.sleep(0.3)
                                    logging.info("execute_vsloader: Deseja visualizar o log detectado. Retornando para o bloco.")
                                    wb_calc = load_workbook(calc_filename)
                                    if "Contas x Fases" not in wb_calc.sheetnames:
                                        ws_contas = wb_calc.create_sheet("Contas x Fases")
                                    else:
                                        ws_contas = wb_calc["Contas x Fases"]
                                    max_r = ws_contas.max_row
                                    for r in range(2, max_r+1):
                                        tfat = ws_contas.cell(row=r, column=1).value
                                        cta = ws_contas.cell(row=r, column=2).value
                                        fse = ws_contas.cell(row=r, column=3).value
                                        if tfat == faturamento_mode.upper() and cta == "311101" and fse == phase_fpp:
                                            ws_contas.cell(row=r, column=4, value="C√°lculo efetuado")
                                            break
                                    wb_calc.save(calc_filename)
                                    continue 
                                elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A √°rea de recibo j√° foi gerada"):
                                    logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    
                                   
                                elif fuzzy_contains(combined_extracted, "ATEN√á√ÉO") and fuzzy_contains(combined_extracted, "Compet√™ncia de Trabalho ser√° alterada"):
                                    logging.info("execute_vsloader: Compet√™ncia de Trabalho ser√° alterada. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("enter")
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Lista de Recibos J√° Emitidos"):
                                    logging.info("execute_vsloader: Lista de recibos j√° emitidos. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    for _ in range(3):
                                        pyautogui.press('enter')
                                        time.sleep(0.3)

                                elif fuzzy_contains(combined_extracted, "Calcular Valores") and fuzzy_contains(combined_extracted, "Midia Inaugural"):
                                    logging.info("execute_vsloader: Calcular valores. Texto identificado: %s", combined_extracted)
                                    break
                                    
                                elif fuzzy_contains(combined_extracted, "<ESC>") or fuzzy_contains(combined_extracted, "Alerta VSSC"):
                                    logging.info("execute_vsloader: Print indica a√ß√£o ESC. Texto identificado: %s", combined_extracted)
                                    pyautogui.moveTo(center_x, center_y)
                                    pyautogui.click()
                                    pyautogui.press("esc")
                                    
                                elif fuzzy_contains(combined_extracted, "Contratos com t√©rmino"):
                                    logging.info("execute_vsloader: Contratos com t√©rmino, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif fuzzy_contains(combined_extracted, "Emiss√£o do Recibo") and fuzzy_contains(combined_extracted, "compet√™ncia"):
                                    logging.info("execute_vsloader: Tela de Emiss√£o de Recibo identificada, desativando valida√ß√£o. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                elif (fuzzy_contains(combined_extracted, "Compet√™ncia de trabalho:") and 
                                    fuzzy_contains(combined_extracted, "Per√≠odo Fechado") and 
                                    fuzzy_contains(combined_extracted, "(Faturamento)")):
                                    logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                                    time.sleep(3)
                                    break
                                else:
                                    logging.info("execute_vsloader: Nenhuma condi√ß√£o modal identificada. Texto identificado: %s", combined_extracted)
                                time.sleep(3)
                            
                            for _ in range(3):
                                pyautogui.moveTo(center_x, center_y)
                                pyautogui.click()
                                pyautogui.press('esc')
                                time.sleep(0.3)

                            break
        for _ in range(5):
            pyautogui.moveTo(center_x, center_y)
            pyautogui.click()
            pyautogui.press('esc')
            time.sleep(0.3)


    #################### CHAMADAS #####################
        login()

        gerar_competencia(tipo_escolhido, from_calculos=True)
        importar_encargos(file_count, txt_files, output_dir, folder)
        executar_calculos() 

    ######################################################
        time.sleep(1)
        delete_all_prints()
        time.sleep(2)
        pyautogui.hotkey('alt', 'F4')
        time.sleep(2)
        pyautogui.moveTo(center_x, center_y)
        pyautogui.click()
        pyautogui.hotkey('alt', 'F4')


    except ElementNotFoundError as e:
        logging.error("Elemento n√£o encontrado: %s" % str(e))
    except TimeoutError as e:
        logging.error("Tempo limite excedido: %s" % str(e))
    except Exception as e:
        logging.error("Erro inesperado: %s" % str(e))

def capture_screenshot(retries: int = 3, delay: float = 0.5) -> str:
    """
    Tenta capturar a tela at√© `retries` vezes caso ocorra OSError.
    Se ainda falhar, propaga o erro.
    """
    for attempt in range(1, retries + 1):
        timestamp = int(time.time() * 1000)
        file_path = os.path.join(prints_folder, f"monitor_screenshot_{timestamp}.png")
        try:
            logging.info(f"[Screenshot] tentativa {attempt}/{retries}: salvando em {file_path}")
            pyautogui.screenshot(file_path)
            return file_path
        except OSError as e:
            logging.warning(f"[Screenshot] falhou na tentativa {attempt}: {e}")
            if attempt < retries:
                time.sleep(delay)
            else:
                logging.error(f"[Screenshot] n√£o foi poss√≠vel capturar ap√≥s {retries} tentativas")
                raise


def delete_all_prints():
    """
    Deleta todos os arquivos na pasta de prints.
    """
    for filename in os.listdir(prints_folder):
        file_path = os.path.join(prints_folder, filename)
        try:
            os.remove(file_path)
            logging.info(f"Print deletado: {file_path}")
        except Exception as e:
            logging.error(f"Erro ao deletar {file_path}: {e}")

def analyze_screenshot(image_path):
    """
    Converte a imagem para Base64 e envia para a API do GPT-4o, retornando o texto extra√≠do.
    """
    if not os.path.exists(image_path):
        logging.error("Nenhuma imagem v√°lida encontrada para an√°lise.")
        return None
    with open(image_path, "rb") as img_file:
        base64_image = base64.b64encode(img_file.read()).decode("utf-8")
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "Identifique a tela em foco no print, normalmente um modal ou um alerta, e me devolva o conte√∫do COMPLETO da tela por escrito, normalmente esse modal, se existir, vai ter t√≠tulo como 'Gerando √Årea de Recibo' ou 'Alerta VSSC', lembrando que a an√°lise deve ser feita principalmente se h√° um modal ou tela no meu aplicativo aberta al√©m da tela principal. Se n√£o houver nenhum modal ou tela, pode-se concluir que o sistema est√° no lobby. Muita aten√ß√£o ao conte√∫do de cada modal se houver mais de um modal detectado. EU PRECISO QUE O TEXTO DA RESPOSTA SEMPRE COMECE COM 'MODAL DETECTADO' OU 'MODAL N√ÉO DETECTADO'",
                        },
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{base64_image}"},
                        },
                    ],
                }
            ],
            request_timeout=60
        )
        extracted_text = response["choices"][0]["message"]["content"].strip()
        return extracted_text
    except Exception as e:
        if "base64" not in str(e):
            logging.error(f"Erro ao analisar imagem: {e}")
        return None


if __name__ == "__main__":
    import sys
    shopping_escolhido = sys.argv[1]
    tipo_escolhido = sys.argv[2]

    # Abrevia√ß√£o dos nomes dos shoppings
    abreviacoes = {
        "Shopping da Ilha": "SDI",
        "Shopping Moxuara": "SMO",
        "Shopping Mestre √Ålvaro": "SMA",
        "Shopping Montserrat": "SMS",
        "Shopping Rio Poty": "SRP",
        "Shopping Metr√≥pole": "SMT"
    }
    shopping_abreviado = abreviacoes.get(shopping_escolhido, shopping_escolhido.replace(' ', '_').upper())

    # Cria (se n√£o existir) a pasta "LOGS" e, dentro dela, a subpasta com a abrevia√ß√£o do shopping
    logs_dir = os.path.join(os.getcwd(), "LOGS", shopping_abreviado)
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)

    # Cria o nome do arquivo de log dentro da pasta de logs
    log_filename = os.path.join(logs_dir, f"Hom_gerar_calculos_{shopping_abreviado}_{tipo_escolhido}.txt")

    # Apaga o log anterior (se existir) antes de criar um novo
    if os.path.exists(log_filename):
        os.remove(log_filename)

    # Remove os handlers j√° configurados (para evitar conflitos)
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    # Configura o logging para escrever no arquivo rec√©m-criado
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format=f"({shopping_escolhido} | {tipo_escolhido}) %(asctime)s %(levelname)s: %(message)s",
        datefmt='%d/%m/%Y %H:%M:%S'
    )

    # Inicia o processo desejado
    execute_vsloader(shopping_escolhido, tipo_escolhido)