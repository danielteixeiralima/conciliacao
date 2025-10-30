# -*- coding: utf-8 -*-

###############################################################################
#                              enviar_email.py                                #
###############################################################################

import ctypes
import pyautogui
import logging
import time
import os
import base64
from anthropic import Anthropic

import pyexcel_xls
from pywinauto import Application
from pywinauto.findwindows import ElementNotFoundError, find_windows
from pywinauto.timings import TimeoutError
import cv2
from pywinauto import Desktop
from datetime import date, timedelta
import calendar
from holidays import Brazil
import openai
import difflib
import pyexcel_xls
from openpyxl import Workbook
from openpyxl import load_workbook
from utils import login, gerar_competencia


pyautogui.FAILSAFE

br_holidays = Brazil()



for w in Desktop(backend="uia").windows():
    logging.info(w.window_text())

openai.api_key = "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"

anthropic = Anthropic(api_key='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')

shopping_fases_tipo2 = {
    "Shopping Mestre Álvaro": {
        "Antecipados": [24, 25, 5, 7],
        "Atípicos": [31, 32, 6, 4],
        "Postecipados": [11, 2, 8, 41]
    },
    "Shopping Montserrat": {
        "Antecipados": [24, 25, 5, 7],
        "Atípicos": [31, 32, 6, 15],
        "Postecipados": [11, 4, 8, 22]
    },
    "Shopping Metrópole": {
        "Antecipados": [12, 13, 7, 18],
        "Atípicos": [31, 32, 6, 11],
        "Postecipados": [8, 9, 2, 36]
    },
    "Shopping Rio Poty": {
        "Antecipados": [12, 13, 7, 18],
        "Atípicos": [31, 32, 6, 11],
        "Postecipados": [8, 9, 2, 23]
    },
    "Shopping da Ilha": {
        "Antecipados": [12, 13, 7, 18], 
        "Atípicos": [31, 32, 6, 11],
        "Postecipados": [8, 9, 2, 36, 37, 44]
    },
    "Shopping Moxuara": {
        "Antecipados": [12, 13, 24, 18],
        "Atípicos": [31, 32, 6, 11],
        "Postecipados": [8, 9, 2, 39]
    }
}



folder_map = {
    "Shopping da Ilha": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_ILHA_HOM",
    "Shopping Mestre Álvaro": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_MESTREALVARO_HOM",
    "Shopping Metrópole": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_METROPOLE_HOM",
    "Shopping Montserrat": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_MONTSERRAT_HOM",
    "Shopping Moxuara": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_MOXUARA_HOM",
    "Shopping Praia da Costa": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_PRAIADACOSTA_HOM",
    "Shopping Rio Poty": r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_TERESINA_HOM"
}

prints_folder = os.path.join(os.getcwd(), "prints")
# A cada screenshot, um nome único será gerado para acumular os prints
SCREENSHOT_PATH = os.path.join(prints_folder, "monitor_screenshot.png")
if not os.path.exists(prints_folder):
    os.makedirs(prints_folder)

IS_SEGURO = False

def fuzzy_contains(text, sub, threshold=0.8):
    """
    Verifica se 'sub' está contido em 'text' com tolerância para pequenas variações.
    Se a correspondência exata não for encontrada, usa uma janela deslizante para comparar
    a similaridade. Retorna True se a similaridade máxima for maior ou igual ao limiar.
    """
    text = text.lower()
    sub = sub.lower()
    if sub in text:
        return True
    max_ratio = 0
    sub_len = len(sub)
    for i in range(len(text) - sub_len + 1):
        segment = text[i:i+sub_len]
        ratio = difflib.SequenceMatcher(None, segment, sub).ratio()
        if ratio > max_ratio:
            max_ratio = ratio
        if max_ratio >= threshold:
            return True
    return False

def build_fase_map(shopping):
    base_y = 33
    step_y = 14
    coords = {}
    missing = missing_phases_map.get(shopping, [])
    real_index = 1
    for fase in range(1, 46):
        if fase in missing:
            continue
        if real_index < 14:
            coords[fase] = base_y + (real_index - 1) * step_y
        else:
            coords[fase] = 215
        real_index += 1
    return coords

missing_phases_map = {
    "Shopping Montserrat": [29, 39, 40, 41, 42, 43, 44, 45],
    "Shopping da Ilha": [3],
    "Shopping Mestre Álvaro": [12, 13, 38, 43, 44, 46, 47, 48, 49],
    "Shopping Metrópole": [3],
    "Shopping Moxuara": [],
    "Shopping Praia da Costa": [27, 42],
    "Shopping Rio Poty": [3, 39, 43, 44, 45, 46, 47, 48, 49]
}

def delete_all_prints():
    """
    Deleta todos os arquivos na pasta de prints.
    """
    for filename in os.listdir(prints_folder):
        file_path = os.path.join(prints_folder, filename)
        try:
            os.remove(file_path)
            logging.info(f"Print deletado: {file_path}")
        except Exception as e:
            logging.error(f"Erro ao deletar {file_path}: {e}")

def capture_screenshot(retries: int = 5, delay: float = 1.0) -> str:
    """
    Captura a tela e salva em 'prints/' com nome único.
    Se falhar, tenta novamente até 'retries' vezes.
    """
    for attempt in range(1, retries + 1):
        try:
            timestamp = int(time.time() * 1000)
            file_path = os.path.join(prints_folder, f"monitor_screenshot_{timestamp}.png")
            logging.info(f"[Screenshot] tentativa {attempt}/{retries}")
            pyautogui.screenshot(file_path)
            return file_path
        except Exception as e:
            logging.warning(f"[Screenshot] falhou na tentativa {attempt}: {e}")
            if attempt < retries:
                time.sleep(delay)
            else:
                # Em vez de travar o programa, apenas loga e retorna None
                logging.error(f"[Screenshot] não foi possível capturar após {retries} tentativas: {e}")
                return None



def analyze_screenshot(image_path):
    """ Converte a imagem para Base64 e envia para a API do GPT-4o, retornando o texto extraído. """
    if not os.path.exists(image_path):
        logging.error("Nenhuma imagem válida encontrada para análise.")
        return None

    with open(image_path, "rb") as img_file:
        base64_image = base64.b64encode(img_file.read()).decode("utf-8")

    while True:  # Loop para garantir que obtenha uma resposta válida
        try:
            response = openai.ChatCompletion.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    "Extraia **exatamente** o texto visível da imagem abaixo, como um OCR literal. "
                                    "Não interprete, não resuma, não complete informações, não descreva. "
                                    "Se houver um modal ou janela em primeiro plano, escreva apenas na primeira linha "
                                    "'MODAL DETECTADO'. Se não houver modal, escreva apenas na primeira linha "
                                    "'MODAL NÃO DETECTADO'. "
                                    "Depois dessa primeira linha, cole fielmente o texto visível do modal/tela, "
                                    "sem comentários adicionais, sem observações ou explicações."
                                ),
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:image/png;base64,{base64_image}"},
                            },
                        ],
                    }
                ],
                request_timeout=60
            )
            extracted_text = response["choices"][0]["message"]["content"].strip()
            if extracted_text:  # Se recebeu uma resposta válida, sai do loop
                return extracted_text

        except Exception as e:
            logging.error(f"Erro ao analisar imagem (tentando novamente): {e}")
            time.sleep(5)

def get_visible_index(shopping, fase_val):
    """
    Retorna quantas fases visíveis existem de 1 até fase_val inclusive,
    desconsiderando as fases ausentes conforme missing_phases_map.
    """
    missing = missing_phases_map.get(shopping, [])
    index = 0
    for f in range(1, fase_val + 1):
        if f not in missing:
            index += 1
    return index

def click_fase_tipo1(shopping, fase):
    """
    Corrige o método de clicar na fase, respeitando as fases ausentes
    também para fases acima de 13.
    """
    if not fase:
        logging.error(f"Fase inválida para {shopping}. Verifique o mapeamento.")
        return
    vi = get_visible_index(shopping, fase)
    if vi <= 14:
        y = 33 + 14 * (vi - 1)
        pyautogui.moveRel(-100, y)
        pyautogui.click()
    else:
        times_to_scroll = vi - 14
        pyautogui.moveRel(2, 215)
        for _ in range(times_to_scroll):
            pyautogui.click()
            time.sleep(0.3)
        pyautogui.moveRel(-100, 0)
        pyautogui.click()

def determine_variant(shopping):
    """
    Determina o variant com base no nome do shopping.
    """
    mapping = {
        "Shopping da Ilha": "SDI",
        "Shopping Mestre Álvaro": "SMA",
        "Shopping Moxuara": "SMO",
        "Shopping Montserrat": "SMS",
        "Shopping Metrópole": "SMT",
        "Shopping Rio Poty": "SRP",
        "Shopping Praia da Costa": "SPC"
    }
    return mapping.get(shopping, "SDI")



def find_and_click_button_with_retry(image_path, max_attempts=10, confidence_range=(0.95, 0.6)):
    try:
        for attempt in range(max_attempts):
            confidence = confidence_range[0] - (attempt * (confidence_range[0] - confidence_range[1]) / max_attempts)
            logging.info("Tentativa %d/%d com confiança %.2f" % (attempt + 1, max_attempts, confidence))
            try:
                button = pyautogui.locateOnScreen(image_path, confidence=confidence)
                if button:
                    x, y = pyautogui.center(button)
                    logging.info("Botão encontrado nas coordenadas: x=%d, y=%d" % (x, y))
                    pyautogui.click(x, y)
                    return True
            except Exception as e:
                logging.error("Erro na tentativa %d: %s" % (attempt + 1, str(e)))
            time.sleep(2)
        logging.error("Botão não encontrado após todas as tentativas")
        return False
    except Exception as e:
        logging.error("Erro ao tentar localizar botão: %s" % str(e))
        return False

def verify_image_visibility(image_path, confidence=0.7, max_retries=10):
    try:
        logging.info("Verificando visibilidade da imagem: %s com confiança %f" % (image_path, confidence))
        for attempt in range(max_retries):
            button_location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if button_location is not None:
                x, y = pyautogui.center(button_location)
                logging.info("Imagem visível nas coordenadas: x=%d, y=%d" % (x, y))
                return button_location
            else:
                logging.error("Tentativa %d/%d: Imagem não encontrada na tela." % (attempt + 1, max_retries))
                time.sleep(2)
        return None
    except Exception as e:
        logging.error("Erro ao verificar visibilidade da imagem: %s" % str(e))
        return None

def find_and_click_button(image_path, confidence=0.95):
    try:
        while True:
            button_location = verify_image_visibility(image_path, confidence=confidence)
            if button_location is not None:
                x, y = pyautogui.center(button_location)
                logging.info("Botão encontrado nas coordenadas: x=%d, y=%d" % (x, y))
                pyautogui.click(x, y)
                break
            else:
                logging.info("Tentando localizar a imagem novamente...")
                time.sleep(2)
    except Exception as e:
        logging.error("Erro ao localizar o botão: %s" % str(e))

# Modificação: A função capture_screenshot agora gera um nome único para cada print
def capture_screenshot():
    """
    Captura a tela e salva na pasta de prints com um nome único baseado em timestamp.
    Retorna o caminho do arquivo gerado.
    """
    timestamp = int(time.time() * 1000)
    file_path = os.path.join(prints_folder, f"monitor_screenshot_{timestamp}.png")
    logging.info(f"Capturando screenshot em {file_path}")
    pyautogui.screenshot(file_path)
    return file_path

def delete_all_prints():
    """
    Deleta todos os arquivos na pasta de prints.
    """
    for filename in os.listdir(prints_folder):
        file_path = os.path.join(prints_folder, filename)
        try:
            os.remove(file_path)
            logging.info(f"Print deletado: {file_path}")
        except Exception as e:
            logging.error(f"Erro ao deletar {file_path}: {e}")

def reset_processing_sheet(shopping, tipo):
    """
    Reseta a planilha de processamento do faturamento para o tipo especificado.
    Limpa as colunas de 'Vencimento' e 'Processamento' na planilha Excel
    que fica em ./LOGS/<abreviação_do_shopping>/{shopping}_boleto.xlsx.
    """
    # assume logs_dir foi definido globalmente como em __main__
    excel_filename = os.path.join(
        logs_dir,
        f"{shopping}_boleto.xlsx"
    )

    if not os.path.exists(excel_filename):
        # Se a planilha não existe, cria uma nova
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan1"
        ws.append(["Antecipado", "", "", "Postecipado", "", "", "Atípico", "", ""])
        ws.append([
            "Fase", "Vencimento", "Processamento",
            "Fase", "Vencimento", "Processamento",
            "Fase", "Vencimento", "Processamento"
        ])
        antecipados   = shopping_fases_tipo2[shopping].get("Antecipados", [])
        postecipados  = shopping_fases_tipo2[shopping].get("Postecipados", [])
        atipicos      = shopping_fases_tipo2[shopping].get("Atípicos", [])
        max_len       = max(len(antecipados), len(postecipados), len(atipicos))
        for i in range(max_len):
            a_fase = antecipados[i] if i < len(antecipados) else ""
            p_fase = postecipados[i] if i < len(postecipados) else ""
            t_fase = atipicos[i] if i < len(atipicos) else ""
            ws.append([a_fase, "", "", p_fase, "", "", t_fase, "", ""])
        wb.save(excel_filename)
        logging.info("Planilha criada e salva como %s", excel_filename)
        return

    # caso já exista, apenas limpa as colunas de Vencimento e Processamento
    wb = load_workbook(excel_filename)
    ws = wb.active

    if tipo == "Antecipados":
        base = 0
    elif tipo == "Postecipados":
        base = 3
    elif tipo == "Atípicos":
        base = 6
    else:
        base = 0

    for row in ws.iter_rows(min_row=3):
        if len(row) >= base + 3:
            row[base + 1].value = ""
            row[base + 2].value = ""

    wb.save(excel_filename)
    logging.info(
        "Planilha de processamento resetada para o tipo '%s': %s",
        tipo, excel_filename
    )





def execute_vsloader(shopping, tipo):
    login()

    gerar_competencia(tipo_escolhido)

    try:
        user32 = ctypes.windll.user32

        def get_foreground_window():
            return user32.GetForegroundWindow()

        def wait_for_stable_focus(prev_handle, max_wait=15):
            start_time = time.time()
            while True:
                if time.time() - start_time > max_wait:
                    logging.info("Tempo máximo de espera por foco estável atingido.")
                    break
                time.sleep(1)
                current_handle = get_foreground_window()
                if current_handle != prev_handle:
                    time.sleep(1.5)
                    second_check = get_foreground_window()
                    if second_check != current_handle:
                        prev_handle = second_check
                        continue
                    break

        def wait_for_focus_change(prev_handle, max_wait=60):
            start_time = time.time()
            while True:
                if time.time() - start_time > max_wait:
                    logging.info("Tempo máximo de espera por mudança de foco atingido.")
                    break
                time.sleep(1)
                current_handle = get_foreground_window()
                if current_handle != prev_handle:
                    break

        fases = shopping_fases_tipo2.get(shopping, {}).get(tipo, [])
        folder = folder_map.get(shopping, r"C:\Program Files\Victor & Schellenberger_FAT_HOM\VSSC_MESTREALVARO_HOM")

        screen_width, screen_height = pyautogui.size()
        center_x = screen_width // 2
        center_y = screen_height // 2

        
        
        
        
       
        
        pyautogui.moveTo(center_x, center_y)
        pyautogui.click()
        for _ in range(5):
            pyautogui.press('esc')
            time.sleep(0.3)
        pyautogui.moveTo(center_x, center_y)
        pyautogui.click()
        current_phase = None
        time.sleep(2)
        

        # dentro de execute_vsloader(...)
        excel_filename = os.path.join(logs_dir, f"{shopping}_boleto.xlsx")


        if not os.path.exists(excel_filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Plan1"
            ws.append(["Antecipado", "", "", "Postecipado", "", "", "Atípico", "", ""])
            ws.append(["Fase", "Vencimento", "Processamento",
                    "Fase", "Vencimento", "Processamento",
                    "Fase", "Vencimento", "Processamento"])
            antecipados = shopping_fases_tipo2[shopping].get("Antecipados", [])
            postecipados = shopping_fases_tipo2[shopping].get("Postecipados", [])
            atipicos = shopping_fases_tipo2[shopping].get("Atípicos", [])
            max_len = max(len(antecipados), len(postecipados), len(atipicos))
            for i in range(max_len):
                a_fase = antecipados[i] if i < len(antecipados) else ""
                p_fase = postecipados[i] if i < len(postecipados) else ""
                t_fase = atipicos[i] if i < len(atipicos) else ""
                ws.append([a_fase, "", "", p_fase, "", "", t_fase, "", ""])
            wb.save(excel_filename)
            logging.info("Planilha criada e salva como %s", excel_filename)
        else:
            wb = load_workbook(excel_filename)
            ws = wb.active

        if tipo == "Antecipados":
            base = 0
        elif tipo == "Postecipados":
            base = 3
        elif tipo == "Atípicos":
            base = 6
        else:
            base = 0


        # logging.info(f"Numero de fasess: {len(fases)}")
        # for row in ws.iter_rows(min_row=3):
        #     row[base+1].value = ""
        #     row[base+2].value = ""

        # pending_fases = []
        # for row in ws.iter_rows(min_row=3):
        #     # Considera a fase pendente somente se o status de processamento estiver em branco (None ou string vazia)
        #     if len(row) > base+2 and row[base].value != "" and (row[base+2].value is None or row[base+2].value == ""):
        #         pending_fases.append(row[base].value)

        # logging.info(f"Número de fases atuais: {len(fases)}")
        # if pending_fases:
        #     fases = pending_fases
        # else:
        #     fases = shopping_fases_tipo2[shopping].get(tipo, [])





        max_reprocessamentos = 4
        reprocess_count = 0
        logging.info(f"Numero de fases: {len(fases)}")
    # while True:
        for fase in fases:
            time.sleep(3)
            pyautogui.hotkey('alt', 's')
            for _ in range(8):
                pyautogui.press('right')
                time.sleep(0.3)
            for _ in range(7):
                pyautogui.press('down')
            pyautogui.press('right')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            prev_handle = get_foreground_window()
            wait_for_focus_change(prev_handle, max_wait=12)
            wait_for_stable_focus(prev_handle)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(350, 510)
            pyautogui.click()
            pyautogui.click()
            time.sleep(5)
            pyautogui.moveTo(center_x, center_y-10)
            pyautogui.click()
            pyautogui.press('enter')
            pyautogui.press('enter')
            prev_handle = get_foreground_window()
            wait_for_focus_change(prev_handle, max_wait=12)
            wait_for_stable_focus(prev_handle)
            logging.info(f"Iniciando processo de geração de boletos.")
            # # Verifica se a fase já foi processada (ou seja, se na coluna "Processamento" o status não está vazio nem já é "Boleto gerado")
            # fase_processada = False
            # for row in ws.iter_rows(min_row=3):
            #     if row[base].value == fase:
            #         if row[base+2].value not in (None, "", "Boleto gerado"):
            #             fase_processada = True
            #         break
            # if fase_processada:
            #     logging.info(f"Fase {fase} já processada, pulando.")
            #     continue

            time.sleep(1)
                    
            # 4) Usa o menu de sistema do Windows para manipular a janela (alt+space -> down -> enter)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')


            # 5) Move o mouse relativo a partir da posição atual e clica duas vezes
            pyautogui.moveRel(180, 90)
            pyautogui.click()
            pyautogui.click()
            
            time.sleep(1)
            pyautogui.hotkey('ctrl', 'f')
            time.sleep(1)
            pyautogui.press('end')
            time.sleep(1)
            for _ in range(5):
                pyautogui.press('backspace')
                time.sleep(0.3)

            # ---------- Substitua apenas a função get_visible_index por esta versão -----------
            def get_visible_index(shopping, fase_val):
                valid_fases = list(range(1, fase_val + 1))
                return len(valid_fases)

            visible_index_current = get_visible_index(shopping, current_phase) if current_phase is not None else 0
            visible_index_target = get_visible_index(shopping, fase)
            diff = visible_index_target - visible_index_current

            # Em vez de mover a seleção com setas, identificamos a fase e digitamos o nome dela
            # Se a fase for menor que 10, formata com dois dígitos (ex: "03 -"); caso contrário, mantém o número
            formatted_phase = f"{fase:02d} -"
            pyautogui.typewrite(formatted_phase)
            pyautogui.press('enter')
            time.sleep(2)
            logging.info(f"Selecionando a fase: {formatted_phase} ")
            while True:
                screenshot1 = capture_screenshot()
                extracted1 = analyze_screenshot(screenshot1)
                time.sleep(2)
                screenshot2 = capture_screenshot()
                extracted2 = analyze_screenshot(screenshot2)

                if not extracted1 or not extracted2:
                    time.sleep(8)
                    continue

                extracted1_l = extracted1.lower() if extracted1 else ""
                extracted2_l = extracted2.lower() if extracted2 else ""

                
                combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                screen_width, screen_height = pyautogui.size()
                center_x = screen_width // 2
                center_y = screen_height // 2
                if fuzzy_contains(combined_extracted, "Alerta VSSC"):
                    logging.info("execute_vsloader: Print indica ação ENTER. Texto identificado: %s", combined_extracted)
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press("enter")
                    break
                elif fuzzy_contains(combined_extracted, "Processando código"):
                    logging.info("execute_vsloader: Processando código. Texto identificado: %s", combined_extracted)
                    time.sleep(10)
                    
                else:
                    break
            time.sleep(1)
            pyautogui.press('down')
            time.sleep(1)
            pyautogui.press('up')
            time.sleep(1)
            logging.info("Escrevendo data de emissão")
            current_phase = fase
            time.sleep(2)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(93, 220)
            pyautogui.click()
            pyautogui.click()
            time.sleep(0.3)
            hoje = date.today()
            data_hoje_formatada = f"{hoje.day:02d}{hoje.month:02d}{hoje.year:04d}"
            pyautogui.write(data_hoje_formatada)
            time.sleep(2)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(93, 195)
            pyautogui.click()
            pyautogui.click()
            logging.info("Escrevendo data de vencimento")

            def is_business_day(d):
                """
                Retorna True se a data 'd' for um dia útil (não sábado, não domingo e não feriado).
                """
                return d.weekday() < 5 and d not in br_holidays

            def adjust_to_business_day(d):
                """
                Se a data 'd' cair em fim de semana ou feriado, avança para o próximo dia útil.
                """
                while not is_business_day(d):
                    d += timedelta(days=1)
                return d

            def get_nth_business_day(year, month, n):
                """
                Retorna o n-ésimo dia útil do mês e ano informados.
                Caso não haja n dias úteis no mês, continua a contagem no mês seguinte.
                """
                count = 0
                day = 1
                last_day = calendar.monthrange(year, month)[1]
                while day <= last_day:
                    current_date = date(year, month, day)
                    if is_business_day(current_date):
                        count += 1
                        if count == n:
                            return current_date
                    day += 1
                d = date(year, month, last_day) + timedelta(days=1)
                while count < n:
                    if is_business_day(d):
                        count += 1
                        if count == n:
                            return d
                    d += timedelta(days=1)
                return d

            def calculate_due_date(shopping, tipo, fase):
                """
                Calcula a data de vencimento de acordo com as regras:
                
                - Para faturamento Atípicos (todos os shoppings, todas as fases):
                    vencimento é no dia 15 do mês vigente (ajustado para o próximo dia útil se cair em fim de semana/feriado).
                
                - Para faturamento Postecipados:
                    * Se o shopping for "Shopping Mestre Álvaro" e a fase for 22:
                        vencimento é fixo no dia 5 do mês seguinte (ajustado se necessário).
                    * Se o shopping for "Shopping Metrópole" e a fase for 36:
                        vencimento é o 10º dia útil do mês seguinte.
                
                - Para faturamento Antecipados:
                    * Se o shopping for "Shopping Mestre Álvaro" (todas as fases):
                        vencimento é fixo no dia 5 do mês seguinte (ajustado se necessário).
                    * Se o shopping for "Shopping Metrópole" e a fase for 18:
                        vencimento é fixo no dia 20 do mês seguinte (ajustado se necessário).
                
                - Caso nenhuma regra específica seja atendida, o vencimento padrão será o primeiro dia útil do mês seguinte.
                """
                today = date.today()
                if tipo == "Atípicos":
                    d = date(today.year, today.month, 15)
                    return adjust_to_business_day(d)
                if today.month == 12:
                    next_year = today.year + 1
                    next_month = 1
                else:
                    next_year = today.year
                    next_month = today.month + 1
                if shopping == "Shopping Mestre Álvaro" and tipo == "Postecipados" and fase == 41:
                    d = date(next_year, next_month, 5)
                    return adjust_to_business_day(d)
                if shopping == "Shopping Metrópole" and tipo == "Postecipados" and fase == 36:
                    return get_nth_business_day(next_year, next_month, 10)
                if shopping == "Shopping Mestre Álvaro" and tipo == "Antecipados":
                    d = date(next_year, next_month, 5)
                    return adjust_to_business_day(d)
                if shopping == "Shopping Metrópole" and tipo == "Antecipados" and fase == 18:
                    d = date(next_year, next_month, 20)
                    return adjust_to_business_day(d)
                d = date(next_year, next_month, 1)
                return adjust_to_business_day(d)

            due_date = calculate_due_date(shopping, tipo, fase)
            data_formatada = f"{due_date.day:02d}{due_date.month:02d}{due_date.year:04d}"
            pyautogui.write(data_formatada)
            logging.info("Clicando para selecionar LUC's")
            time.sleep(2)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(-80, 195)
            pyautogui.click()
            pyautogui.click()
            logging.info("Clicando para selecionar LUC's")

            prev_handle = get_foreground_window()
            wait_for_focus_change(prev_handle)
            wait_for_stable_focus(prev_handle)
            time.sleep(2)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(-220, 365)
            pyautogui.click()
            pyautogui.click()
            logging.info("Selecionando todas as empresas")
            time.sleep(0.5)
            for _ in range(3):
                pyautogui.press('tab')
                time.sleep(0.1)
            pyautogui.press('enter')
            time.sleep(1)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(180, 300)
            pyautogui.click()
            pyautogui.click()
            logging.info("Clicando no formato do email gerado")
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(0.5)
            for _ in range(6):
                pyautogui.press('tab')
                time.sleep(0.3)
            pyautogui.press('enter')
            time.sleep(0.3)

            pyautogui.press('enter')
            time.sleep(0.3)
            pyautogui.press('enter')
            time.sleep(2)
            
            pyautogui.press('tab')
            time.sleep(2)
            pyautogui.hotkey('alt', 'space')
            time.sleep(0.3)
            pyautogui.press('down')
            time.sleep(0.3)
            pyautogui.press('enter')
            pyautogui.moveRel(-70, 111)
            pyautogui.click()
            pyautogui.click()
            logging.info("Clicando para gerar o boleto")
            pyautogui.press('tab')
            pyautogui.press('enter')
            time.sleep(0.3)
            pyautogui.press('enter')

            logging.info(f"Gerando boleto.")

            global IS_SEGURO
            while True:
                screenshot1 = capture_screenshot()
                extracted1 = analyze_screenshot(screenshot1)
                time.sleep(2)
                screenshot2 = capture_screenshot()
                extracted2 = analyze_screenshot(screenshot2)
                if extracted1 is None and extracted2 is None:
                    time.sleep(3)
                    continue
                if extracted1 is None and extracted2 is None:
                    time.sleep(3)
                    continue
                extracted1_l = extracted1.lower() if extracted1 else ""
                extracted2_l = extracted2.lower() if extracted2 else ""

                
                combined_extracted = (extracted1 + " " + extracted2) if extracted1 and extracted2 else (extracted1 or extracted2)
                screen_width, screen_height = pyautogui.size()
                center_x = screen_width // 2
                center_y = screen_height // 2

                if fuzzy_contains(combined_extracted, "Processando código"):
                    logging.info("execute_vsloader: Processando código. Texto identificado: %s", combined_extracted)
                    time.sleep(15)
                elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "Relatório gerado"):
                    logging.info("execute_vsloader: A área de recibo já foi gerada. Texto identificado: %s", combined_extracted)
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press("esc")
                    if tipo == "Antecipados":
                        base = 0
                    elif tipo == "Postecipados":
                        base = 3
                    elif tipo == "Atípicos":
                        base = 6
                    else:
                        base = 0

                    updated = False
                    for row in ws.iter_rows(min_row=3):
                        logging.info("Verificando linha: Fase=%s, Vencimento=%s, Processamento=%s", row[base].value, row[base+1].value, row[base+2].value)
                        if row[base].value == fase:
                            row[base+2].value = "Boleto gerado"
                            updated = True
                            logging.info("Atualizado boleto no Excel para fase %s: %s", fase, row[base+2].value)
                            break
                    if not updated:
                        logging.warning("Nenhuma linha encontrada no Excel para atualizar para fase %s", fase)
                    wb.save(excel_filename)
                    logging.info("Planilha atualizada e salva como %s", excel_filename)
                
                elif fuzzy_contains(combined_extracted, "Alerta VSSC") and fuzzy_contains(combined_extracted, "A área de recibo já foi gerada"):
                    logging.info("execute_vsloader: A área de recibo já foi gerada. Texto identificado: %s", combined_extracted)
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    time.sleep(0.5)
                    pyautogui.press("enter")
                
                elif fuzzy_contains(combined_extracted, "ATENÇÃO") and fuzzy_contains(combined_extracted, "Competência de Trabalho será alterada"):
                    logging.info("execute_vsloader: Competência de Trabalho será alterada. Texto identificado: %s", combined_extracted)
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    pyautogui.press("enter")
                    time.sleep(3)
                    break
                elif fuzzy_contains(combined_extracted, "Lista de Recibos Já Emitidos"):
                    logging.info("execute_vsloader: Lista de Recibos Já Emitidos. Texto identificado: %s", combined_extracted)
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    for _ in range(3):
                        pyautogui.press('tab')
                        time.sleep(0.3)
                    pyautogui.press('enter')
                    
                elif fuzzy_contains(combined_extracted, "Pressione <ESC>"):
                    logging.info("execute_vsloader: Print indica ação ESC. Texto identificado: %s", combined_extracted) 
                    # pyautogui.moveTo(center_x, center_y)
                    # pyautogui.click()
                    for _ in range(5):
                        pyautogui.press('esc')
                        time.sleep(0.5)
                    break

                elif fuzzy_contains(combined_extracted, "Alerta VSSC"):
                    logging.info("execute_vsloader: Alerta VSSC. Texto identificado: %s", combined_extracted) 
                    pyautogui.moveTo(center_x, center_y)
                    pyautogui.click()
                    time.sleep(0.5)
                    pyautogui.press("esc")
                
                elif fuzzy_contains(combined_extracted, "Contratos com término") and not fuzzy_contains(combined_extracted, "Contrato não está marcado"):
                    logging.info("execute_vsloader: Contratos com término, desativando validação. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                    break
                # elif fuzzy_contains(combined_extracted, "Emissão do Recibo") and fuzzy_contains(combined_extracted, "competência"):
                #     logging.info("execute_vsloader: Tela de Emissão de Recibo identificada, desativando validação. Texto identificado: %s", combined_extracted)
                #     time.sleep(4)
                #     break
                elif fuzzy_contains(combined_extracted, "Emitindo Recibos"):
                    logging.info("execute_vsloader: Emitindo Recibos, desativando validação. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                
                elif (fuzzy_contains(combined_extracted, "Competência de trabalho:") and 
                    fuzzy_contains(combined_extracted, "Período Fechado") and 
                    fuzzy_contains(combined_extracted, "(Faturamento)")):
                    logging.info("execute_vsloader: Lobby identificado. Texto identificado: %s", combined_extracted)
                    time.sleep(3)
                    break
                else:
                    logging.info("execute_vsloader: Nenhuma condição modal identificada. Texto identificado: %s", combined_extracted)
                    time.sleep(10)
                time.sleep(3)
            
            for _ in range(3):
                pyautogui.moveTo(center_x, center_y)
                pyautogui.click()
                pyautogui.press('esc')
                time.sleep(0.3)
            
            # pending_fases = []
            # for row in ws.iter_rows(min_row=3):
            #     if len(row) > base+1 and row[base].value != "" and row[base+2].value != "Boleto gerado":
            #         pending_fases.append(row[base].value)

            # if pending_fases:
            #     reprocess_count += 1
            #     if reprocess_count > max_reprocessamentos:
            #         logging.warning("Número máximo de reprocessamentos atingido. Encerrando loop.")
            #         break
            #     pyautogui.moveTo(center_x, center_y)
            #     pyautogui.click()
            #     current_phase = None
            #     time.sleep(2)
            #     pyautogui.hotkey('alt', 's')
            #     for _ in range(8):
            #         pyautogui.press('right')
            #         time.sleep(0.3)
            #     for _ in range(7):
            #         pyautogui.press('down')
            #     pyautogui.press('right')
            #     time.sleep(0.3)
            #     pyautogui.press('down')
            #     time.sleep(0.3)
            #     pyautogui.press('enter')
            #     prev_handle = get_foreground_window()
            #     wait_for_focus_change(prev_handle)
            #     wait_for_stable_focus(prev_handle)
            #     pyautogui.hotkey('alt', 'space')
            #     time.sleep(0.3)
            #     pyautogui.press('down')
            #     time.sleep(0.3)
            #     pyautogui.press('enter')
            #     pyautogui.moveRel(350, 510)
            #     pyautogui.click()
            #     pyautogui.click()
            #     time.sleep(5)
            #     pyautogui.moveTo(center_x, center_y-10)
            #     pyautogui.click()
            #     pyautogui.press('enter')
            #     prev_handle = get_foreground_window()
            #     wait_for_focus_change(prev_handle)
            #     wait_for_stable_focus(prev_handle)
            #     if pending_fases:
            #         logging.info("Reiniciando processo de geração de boletos para as fases pendentes.")
            #         fases = pending_fases
            #         continue
            #     else:
            #         break
            # else:
            #     break
        for _ in range(5):
            pyautogui.press('esc')
            time.sleep(0.3)            
        wb.save(excel_filename)

        logging.info("Planilha salva com sucesso como %s_boleto.xlsx", shopping)

#################### CHAMADAS #####################

        
   #########################################


                # Verifica se há alguma fase com processamento vazio (pendente)
                # Verifica se há alguma fase com processamento pendente (ou seja, com a coluna "Processamento" vazia)
        
        delete_all_prints()
        time.sleep(2)
        pyautogui.hotkey('alt', 'F4')
        time.sleep(2)
        pyautogui.moveTo(center_x, center_y)
        pyautogui.click()
        pyautogui.hotkey('alt', 'F4')


        

    except ElementNotFoundError as e:
        logging.error(f"Elemento não encontrado: {e}")
    except TimeoutError as e:
        logging.error(f"Tempo limite excedido: {e}")
    except Exception as e:
        logging.error(f"Erro inesperado: {e}")



if __name__ == "__main__":
    import sys
    import os
    shopping_escolhido = sys.argv[1]
    tipo_escolhido    = sys.argv[2]

    # Abreviação dos nomes dos shoppings
    abreviacoes = {
        "Shopping da Ilha": "SDI",
        "Shopping Moxuara": "SMO",
        "Shopping Mestre Álvaro": "SMA",
        "Shopping Montserrat": "SMS",
        "Shopping Rio Poty": "SRP",
        "Shopping Metrópole": "SMT"
    }
    shopping_abreviado = abreviacoes.get(
        shopping_escolhido,
        shopping_escolhido.replace(" ", "_").upper()
    )

    # --- cria a pasta de logs dentro de ./LOGS/<abreviação> ---
    logs_dir = os.path.join(os.getcwd(), "LOGS", shopping_abreviado)
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)

    # nome do arquivo de log dentro de LOGS/<abreviação>/
    log_filename = os.path.join(
        logs_dir,
        f"Gerar_boletos_{shopping_abreviado}_{tipo_escolhido}.txt"
    )

    # remove handlers antigos e configura logging para escrever no novo arquivo
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format=f"({shopping_escolhido} | {tipo_escolhido}) %(asctime)s %(levelname)s: %(message)s",
        datefmt="%d/%m/%Y %H:%M:%S"
    )

    reset_processing_sheet(shopping_escolhido, tipo_escolhido)
    execute_vsloader(shopping_escolhido, tipo_escolhido)

